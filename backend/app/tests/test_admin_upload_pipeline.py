"""
Teste de Integração - Pipeline Completo de Upload Admin
========================================================

Valida:
- Geração de arquivo Excel teste
- Upload via endpoint admin
- Snapshot extraction
- Field detection
- Storage persistence
- DB registration
"""

import pytest
import io
import tempfile
import json
import gzip
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "app"))

from app.services.template_snapshot import TemplateSnapshotService, validate_snapshot
from app.services.fillable_detector import FillableAreaDetector
from app.services.template_storage import TemplateStorageService
from app.services.template_registry import TemplateRegistry
from app.models.template_definition import TemplateDefinition, FillableField, Base


@pytest.fixture
def in_memory_db():
    """Cria DB SQLite em memória"""
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    
    Session = sessionmaker(bind=engine)
    session = Session()
    
    yield session
    
    session.close()


@pytest.fixture
def temp_storage():
    """Cria diretório temporário para storage"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir


@pytest.fixture
def fcj_template_excel():
    """Cria template Excel FCJ de teste"""
    wb = Workbook()
    
    # ============ ICP Sheet ============
    ws_icp = wb.active
    ws_icp.title = "ICP"
    
    # Títulos
    ws_icp["A1"] = "IDEAL CUSTOMER PROFILE"
    ws_icp["A1"].font = Font(bold=True, size=14)
    ws_icp["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_icp["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    
    # Campos ICP
    ws_icp["A3"] = "Nome da Empresa:"
    ws_icp.merge_cells("B3:D3")
    ws_icp["B3"].value = None
    
    ws_icp["A4"] = "Setor:"
    ws_icp.merge_cells("B4:D4")
    
    dv_setor = DataValidation(type="list", formula1='"Tech,Finance,Healthcare,Retail"')
    ws_icp.add_data_validation(dv_setor)
    dv_setor.add("B4")
    ws_icp["B4"].value = None
    
    ws_icp["A5"] = "Tamanho da Empresa (ARR):"
    ws_icp.merge_cells("B5:D5")
    ws_icp["B5"].number_format = "$#,##0.00"
    ws_icp["B5"].value = None
    
    # ============ Persona Sheet ============
    ws_persona = wb.create_sheet("Persona")
    
    ws_persona["A1"] = "PERSONA"
    ws_persona["A1"].font = Font(bold=True, size=14)
    ws_persona["A1"].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws_persona["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    
    ws_persona["A3"] = "Nome:"
    ws_persona.merge_cells("B3:D3")
    ws_persona["B3"].value = None
    
    ws_persona["A4"] = "Cargo:"
    ws_persona.merge_cells("B4:D4")
    ws_persona["B4"].value = None
    
    ws_persona["A5"] = "Responsabilidades:"
    ws_persona.merge_cells("A6:D8")
    ws_persona["A6"].alignment = Alignment(wrap_text=True)
    ws_persona["A6"].value = None
    
    # ============ SWOT Sheet ============
    ws_swot = wb.create_sheet("SWOT")
    
    ws_swot["A1"] = "SWOT ANALYSIS"
    ws_swot["A1"].font = Font(bold=True, size=14)
    ws_swot["A1"].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws_swot["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    
    # 2x2 SWOT grid
    ws_swot["A3"] = "Strengths"
    ws_swot["C3"] = "Weaknesses"
    ws_swot["A5"] = "Opportunities"
    ws_swot["C5"] = "Threats"
    
    # Campos SWOT
    ws_swot.merge_cells("A4:B4")
    ws_swot["A4"].alignment = Alignment(wrap_text=True)
    
    ws_swot.merge_cells("C4:D4")
    ws_swot["C4"].alignment = Alignment(wrap_text=True)
    
    ws_swot.merge_cells("A6:B6")
    ws_swot["A6"].alignment = Alignment(wrap_text=True)
    
    ws_swot.merge_cells("C6:D6")
    ws_swot["C6"].alignment = Alignment(wrap_text=True)
    
    # Salvar em bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


class TestAdminUploadPipeline:
    """Testes de integração do pipeline de upload"""
    
    def test_full_pipeline(self, fcj_template_excel, in_memory_db, temp_storage):
        """Testa pipeline completo de upload"""
        
        # 1. SNAPSHOT EXTRACTION
        print("\n📊 1. SNAPSHOT EXTRACTION")
        snapshot_service = TemplateSnapshotService()
        snapshot, assets = snapshot_service.extract(fcj_template_excel)
        
        assert snapshot["schema_version"] == "2.0"
        assert len(snapshot["sheets"]) == 3
        print(f"   ✓ Snapshot extraído: {len(snapshot['sheets'])} sheets")
        
        # 2. SNAPSHOT VALIDATION
        print("\n✅ 2. SNAPSHOT VALIDATION")
        validation = validate_snapshot(snapshot)
        assert validation["valid"] == True
        print(f"   ✓ Snapshot válido")
        
        # 3. FILLABLE DETECTION
        print("\n🔍 3. FILLABLE DETECTION")
        detector = FillableAreaDetector()
        candidates = detector.detect(snapshot)
        
        assert len(candidates) > 0
        print(f"   ✓ Detectados {len(candidates)} campos")
        
        # 4. STORAGE PERSISTENCE
        print("\n💾 4. STORAGE PERSISTENCE")
        storage = TemplateStorageService(base_path=temp_storage)
        registry = TemplateRegistry()
        
        file_hash = registry.compute_file_hash(fcj_template_excel)
        template_key = registry.compute_template_key("test_fcj.xlsx", "Q1")
        
        save_result = storage.save(
            file_name="test_fcj.xlsx",
            file_bytes=fcj_template_excel,
            snapshot_dict=snapshot,
            assets=assets,
            template_key=template_key,
            cycle="Q1"
        )
        
        assert save_result["hash"] == file_hash
        assert Path(save_result["paths"]["original_path"]).exists()
        print(f"   ✓ Storage salvo")
        
        # 5. DB REGISTRATION
        print("\n🗄️ 5. DB REGISTRATION")
        
        fields_payload = [c.to_dict(template_id="pending") for c in candidates]
        stats = registry.compute_stats(snapshot, fields_payload)
        
        td = registry.upsert_template_definition(
            db=in_memory_db,
            template_key=template_key,
            cycle="Q1",
            file_hash=file_hash,
            original_path=save_result["paths"]["original_path"],
            snapshot_path=save_result["paths"]["snapshot_path"],
            assets_manifest_path=save_result["paths"].get("assets_manifest_path"),
            stats=stats
        )
        
        assert td.id is not None
        print(f"   ✓ Template registrado (id={td.id})")
        
        # 6. FIELDS REGISTRATION
        print("\n📝 6. FIELDS REGISTRATION")
        
        fields_final = [c.to_dict(template_id=str(td.id)) for c in candidates]
        registry.replace_fields_for_template(
            db=in_memory_db,
            template_id=td.id,
            fields=fields_final
        )
        
        in_memory_db.flush()
        
        stored_fields = in_memory_db.query(FillableField).filter_by(
            template_id=td.id
        ).all()
        
        assert len(stored_fields) == len(candidates)
        print(f"   ✓ Fields registrados: {len(stored_fields)}")
        
        print("\n✅ PIPELINE COMPLETO VALIDADO\n")
