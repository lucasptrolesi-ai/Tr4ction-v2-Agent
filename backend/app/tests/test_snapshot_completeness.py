"""
Testes do Template Snapshot Service
===================================

Valida extração completa e sem perda de informação
"""

import pytest
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "app"))

from app.services.template_snapshot import (
    TemplateSnapshotService,
    validate_snapshot,
    SnapshotValidationError
)


class TestSnapshotCompleteness:
    """Testes de completude do snapshot"""
    
    def test_extract_basic_workbook(self):
        """Testa extração de workbook básico"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        ws["A1"] = "Header"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Value"
        ws.merge_cells("B2:C2")
        ws["B2"] = "Merged"
        
        # Salvar em bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        file_bytes = buffer.getvalue()
        
        # Extrair
        service = TemplateSnapshotService()
        snapshot, assets = service.extract(file_bytes)
        
        # Validações
        assert snapshot["schema_version"] == "2.0"
        assert len(snapshot["sheets"]) == 1
        
        sheet = snapshot["sheets"][0]
        assert sheet["name"] == "Test Sheet"
        assert "cells" in sheet
        assert "merged_cells" in sheet
        assert "B2:C2" in sheet["merged_cells"]
        
        # Validar estilos
        cells = sheet["cells"]
        header_cell = next(c for c in cells if c["coordinate"] == "A1")
        assert header_cell["style"]["font"]["bold"] == True
        assert header_cell["style"]["font"]["size"] == 14
    
    def test_extract_with_validations(self):
        """Testa extração de data validations"""
        from openpyxl.worksheet.datavalidation import DataValidation
        
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Select"
        
        dv = DataValidation(type="list", formula1='"Option1,Option2,Option3"')
        ws.add_data_validation(dv)
        dv.add("A2:A10")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        file_bytes = buffer.getvalue()
        
        service = TemplateSnapshotService()
        snapshot, assets = service.extract(file_bytes)
        
        sheet = snapshot["sheets"][0]
        assert len(sheet["data_validations"]) > 0
        
        validation = sheet["data_validations"][0]
        assert validation["type"] == "list"
        assert "Option1" in validation["formula1"]
    
    def test_extract_with_merged_and_styles(self):
        """Testa extração de merged cells com estilos"""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:B1")
        ws["A1"] = "Merged Title"
        ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        file_bytes = buffer.getvalue()
        
        service = TemplateSnapshotService()
        snapshot, assets = service.extract(file_bytes)
        
        sheet = snapshot["sheets"][0]
        assert "A1:B1" in sheet["merged_cells"]
        
        cell = next(c for c in sheet["cells"] if c["coordinate"] == "A1")
        assert cell["style"]["fill"]["fgColor"] is not None
        assert cell["style"]["alignment"]["horizontal"] == "center"
    
    def test_snapshot_validation_pass(self):
        """Testa validação de snapshot válido"""
        snapshot = {
            "schema_version": "2.0",
            "sheets": [
                {
                    "name": "Sheet1",
                    "cells": [
                        {
                            "coordinate": "A1",
                            "value": "Test",
                            "style": {
                                "font": {},
                                "fill": {},
                                "border": {},
                                "alignment": {},
                                "protection": {}
                            }
                        }
                    ],
                    "merged_cells": [],
                    "row_dimensions": [],
                    "column_dimensions": [],
                    "data_validations": [],
                    "conditional_formatting": [],
                    "tables": [],
                    "images": [],
                }
            ]
        }
        
        result = validate_snapshot(snapshot)
        assert result["valid"] == True
        assert len(result["errors"]) == 0
    
    def test_snapshot_validation_fail_missing_components(self):
        """Testa validação de snapshot inválido"""
        snapshot = {
            "schema_version": "2.0",
            "sheets": [
                {
                    "name": "Sheet1",
                    "cells": [],
                    # Faltando componentes obrigatórios
                }
            ]
        }
        
        result = validate_snapshot(snapshot)
        assert result["valid"] == False
        assert len(result["errors"]) > 0


class TestSnapshotStats:
    """Testes de estatísticas do snapshot"""
    
    def test_stats_computation(self):
        """Testa computação de stats"""
        wb = Workbook()
        ws = wb.active
        
        # Adicionar conteúdo
        for i in range(1, 11):
            ws[f"A{i}"] = f"Value {i}"
        
        ws.merge_cells("B1:C1")
        ws["B1"] = "Merged"
        
        buffer = io.BytesIO()
        wb.save(buffer)
        file_bytes = buffer.getvalue()
        
        service = TemplateSnapshotService()
        snapshot, assets = service.extract(file_bytes)
        
        result = validate_snapshot(snapshot)
        stats = result["stats"]
        
        assert stats["sheets_count"] == 1
        assert stats["total_cells"] >= 10
        assert stats["total_merged"] >= 1


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
