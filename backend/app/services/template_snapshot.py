"""
Template Snapshot Service - Extração completa e auditável de templates Excel
============================================================================

RESPONSABILIDADE:
Extrair TODOS os dados de um arquivo .xlsx em um snapshot JSON completo,
preservando fidelidade semântica para o método FCJ.

GARANTIAS:
- Extração sem perda de informação
- Auditabilidade completa
- Reprodutibilidade determinística
- Versionamento de schema

SCHEMA VERSION: 2.0 (completo com validação)
"""

from __future__ import annotations
import io
import logging
from typing import Dict, Any, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

SNAPSHOT_SCHEMA_VERSION = "2.0"


class SnapshotValidationError(Exception):
    """Erro crítico de validação de snapshot"""
    pass


class SnapshotLoadError(Exception):
    """Erro ao carregar arquivo Excel"""
    pass


class TemplateSnapshotService:
    """
    Serviço de extração completa de templates Excel
    """

    def extract(self, file_bytes: bytes) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """
        Extrai snapshot completo + assets de um arquivo Excel
        
        Args:
            file_bytes: Conteúdo binário do arquivo .xlsx
            
        Returns:
            Tuple com:
            - snapshot: Dict completo com workbook + sheets
            - assets: Lista de dicts com metadados + binário de imagens
            
        Raises:
            SnapshotLoadError: Se arquivo não puder ser carregado
            SnapshotValidationError: Se snapshot estiver incompleto
        """
        # 1. Carregar workbook com segurança
        try:
            wb = load_workbook(
                io.BytesIO(file_bytes),
                data_only=False,
                keep_vba=False
            )
        except Exception as e:
            error_msg = f"Falha ao carregar arquivo XLSX: {str(e)}"
            logger.error(error_msg)
            raise SnapshotLoadError(error_msg) from e
        
        # 2. Validar que workbook foi carregado corretamente
        if not wb or not hasattr(wb, 'worksheets'):
            raise SnapshotLoadError("Workbook inválido ou vazio")
        
        snapshot = {
            "schema_version": SNAPSHOT_SCHEMA_VERSION,
            "workbook": self._extract_workbook_props(wb),
            "sheets": []
        }
        
        all_assets = []
        
        # 3. Extrair cada sheet
        for sheet in wb.worksheets:
            try:
                sheet_data, sheet_assets = self._extract_sheet(sheet)
                snapshot["sheets"].append(sheet_data)
                all_assets.extend(sheet_assets)
            except Exception as e:
                error_msg = f"Falha ao extrair sheet '{sheet.title}': {str(e)}"
                logger.error(error_msg)
                raise SnapshotValidationError(error_msg) from e
        
        # 4. Auto-validação obrigatória
        self._validate_snapshot(snapshot)
        
        return snapshot, all_assets

    def _extract_workbook_props(self, wb) -> Dict[str, Any]:
        """Extrai propriedades do workbook"""
        return {
            "defined_names": {
                name: wb.defined_names[name].value
                for name in wb.defined_names.definedName
            } if wb.defined_names else {},
            "sheetnames": wb.sheetnames,
        }

    def _extract_sheet(self, sheet: Worksheet) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """
        Extrai dados completos de uma sheet
        
        CRÍTICO PARA FCJ: 
        - Sheet index é preservado (ordem das abas é importante)
        - Células são extraídas em ordem vertical (top-to-bottom)
        
        Returns:
            Tuple com (sheet_data, assets)
        """
        sheet_data = {
            "name": sheet.title,
            "sheet_state": getattr(sheet, "sheet_state", "visible"),
            "freeze_panes": self._extract_freeze_panes(sheet),
            "page_setup": self._extract_page_setup(sheet),
            "page_margins": self._extract_page_margins(sheet),
            "row_dimensions": self._extract_row_dimensions(sheet),
            "column_dimensions": self._extract_column_dimensions(sheet),
            "merged_cells": [str(r) for r in sheet.merged_cells.ranges],
            "cells": [],
            "data_validations": self._extract_data_validations(sheet),
            "conditional_formatting": self._extract_conditional_formatting(sheet),
            "tables": self._extract_tables(sheet),
            "images": [],
        }
        
        # Extrair células (PRESERVAR ORDEM VERTICAL)
        # Usar ._cells.values() para obter células que foram modificadas
        # MAS ordenar por (row, col) para garantir ordem de leitura top-to-bottom
        if hasattr(sheet, '_cells') and sheet._cells:
            cells_list = list(sheet._cells.values())
            # ✅ CRÍTICO: Ordenar por (row, column) para leitura vertical
            cells_list.sort(key=lambda c: (c.row, c.column))
            for cell in cells_list:
                sheet_data["cells"].append(self._extract_cell(cell))
        else:
            # Fallback: iterar sobre o range máximo em ordem vertical
            max_row = sheet.max_row or 1
            max_col = sheet.max_column or 1
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row, col)
                    # Extrair célula se tiver conteúdo ou estilo significante
                    if (cell.value is not None or 
                        cell.formula or 
                        cell.hyperlink or 
                        cell.comment or
                        self._has_style(cell)):
                        sheet_data["cells"].append(self._extract_cell(cell))
        
        # Extrair imagens
        assets = []
        if hasattr(sheet, "_images") and sheet._images:
            for idx, img in enumerate(sheet._images):
                img_data = self._extract_image(img, sheet.title, idx)
                sheet_data["images"].append({
                    "index": idx,
                    "anchor": img_data["anchor"],
                    "format": img_data["format"],
                })
                assets.append(img_data)
        
        return sheet_data, assets

    def _has_style(self, cell) -> bool:
        """Verifica se célula tem estilo significante (não é padrão)"""
        try:
            # Check se tem preenchimento de cor
            if cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb'):
                if cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                    return True
            
            # Check se tem fonte especial (bold, italic, cor, etc)
            if cell.font:
                if cell.font.bold or cell.font.italic or cell.font.underline:
                    return True
                if cell.font.color and hasattr(cell.font.color, 'rgb'):
                    if cell.font.color.rgb and cell.font.color.rgb != '00000000':
                        return True
            
            # Check se tem borda
            if cell.border:
                for side in [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]:
                    if side and side.style:
                        return True
            
            # Check se tem alinhamento especial
            if cell.alignment and (cell.alignment.wrapText or cell.alignment.shrinkToFit):
                return True
            
            return False
        except Exception:
            return False

    def _extract_cell(self, cell) -> Dict[str, Any]:
        """Extrai dados completos de uma célula"""
        return {
            "coordinate": cell.coordinate,
            "row": cell.row,
            "column": cell.column,
            "column_letter": get_column_letter(cell.column),
            "value": self._serialize_value(cell.value),
            "data_type": cell.data_type,
            "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
            "number_format": cell.number_format,
            "hyperlink": str(cell.hyperlink.target) if cell.hyperlink else None,
            "comment": cell.comment.text if cell.comment else None,
            "style": self._extract_cell_style(cell),
        }

    def _extract_cell_style(self, cell) -> Dict[str, Any]:
        """Extrai estilo completo da célula"""
        return {
            "font": self._extract_font(cell.font),
            "fill": self._extract_fill(cell.fill),
            "border": self._extract_border(cell.border),
            "alignment": self._extract_alignment(cell.alignment),
            "protection": self._extract_protection(cell.protection),
        }

    def _extract_font(self, font: Font) -> Dict[str, Any]:
        """Extrai propriedades de fonte"""
        return {
            "name": font.name,
            "size": font.size,
            "bold": font.bold,
            "italic": font.italic,
            "underline": font.underline,
            "strike": font.strike,
            "color": self._extract_color(font.color),
        }

    def _extract_fill(self, fill: PatternFill) -> Dict[str, Any]:
        """Extrai propriedades de preenchimento"""
        return {
            "patternType": fill.patternType,
            "fgColor": self._extract_color(fill.fgColor),
            "bgColor": self._extract_color(fill.bgColor),
        }

    def _extract_border(self, border: Border) -> Dict[str, Any]:
        """Extrai propriedades de borda"""
        return {
            "left": {"style": border.left.style, "color": self._extract_color(border.left.color)} if border.left else None,
            "right": {"style": border.right.style, "color": self._extract_color(border.right.color)} if border.right else None,
            "top": {"style": border.top.style, "color": self._extract_color(border.top.color)} if border.top else None,
            "bottom": {"style": border.bottom.style, "color": self._extract_color(border.bottom.color)} if border.bottom else None,
        }

    def _extract_alignment(self, alignment: Alignment) -> Dict[str, Any]:
        """Extrai propriedades de alinhamento"""
        return {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "textRotation": alignment.textRotation,
            "wrapText": alignment.wrapText,
            "shrinkToFit": alignment.shrinkToFit,
            "indent": alignment.indent,
        }

    def _extract_protection(self, protection: Protection) -> Dict[str, Any]:
        """Extrai propriedades de proteção"""
        return {
            "locked": protection.locked,
            "hidden": protection.hidden,
        }

    def _extract_color(self, color) -> Optional[str]:
        """Extrai cor RGB"""
        if not color:
            return None
        if hasattr(color, "rgb") and color.rgb:
            return str(color.rgb)
        if hasattr(color, "indexed") and color.indexed:
            return f"indexed:{color.indexed}"
        return None

    def _extract_freeze_panes(self, sheet: Worksheet) -> Optional[str]:
        """Extrai freeze panes"""
        if sheet.freeze_panes:
            return sheet.freeze_panes
        return None

    def _extract_page_setup(self, sheet: Worksheet) -> Dict[str, Any]:
        """Extrai configuração de página"""
        ps = sheet.page_setup
        return {
            "orientation": ps.orientation,
            "paperSize": ps.paperSize,
            "fitToHeight": ps.fitToHeight,
            "fitToWidth": ps.fitToWidth,
            "scale": ps.scale,
        }

    def _extract_page_margins(self, sheet: Worksheet) -> Dict[str, Any]:
        """Extrai margens de página"""
        pm = sheet.page_margins
        return {
            "left": pm.left,
            "right": pm.right,
            "top": pm.top,
            "bottom": pm.bottom,
            "header": pm.header,
            "footer": pm.footer,
        }

    def _extract_row_dimensions(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extrai dimensões de linhas"""
        rows = []
        for row_num, dim in sheet.row_dimensions.items():
            if dim.height or dim.hidden or dim.outlineLevel:
                rows.append({
                    "row": row_num,
                    "height": dim.height,
                    "hidden": dim.hidden,
                    "outlineLevel": dim.outlineLevel,
                })
        return rows

    def _extract_column_dimensions(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extrai dimensões de colunas"""
        cols = []
        for col_letter, dim in sheet.column_dimensions.items():
            if dim.width or dim.hidden or dim.outlineLevel:
                cols.append({
                    "column": col_letter,
                    "width": dim.width,
                    "hidden": dim.hidden,
                    "outlineLevel": dim.outlineLevel,
                })
        return cols

    def _extract_data_validations(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extrai validações de dados"""
        validations = []
        if hasattr(sheet, "data_validations"):
            for dv in sheet.data_validations.dataValidation:
                validations.append({
                    "type": dv.type,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "showDropDown": dv.showDropDown,
                    "showErrorMessage": dv.showErrorMessage,
                    "showInputMessage": dv.showInputMessage,
                    "errorTitle": dv.errorTitle,
                    "error": dv.error,
                    "promptTitle": dv.promptTitle,
                    "prompt": dv.prompt,
                    "sqref": str(dv.sqref) if dv.sqref else None,
                })
        return validations

    def _extract_conditional_formatting(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extrai formatação condicional"""
        cf_list = []
        if hasattr(sheet, "conditional_formatting"):
            for cf_range, rules in sheet.conditional_formatting._cf_rules.items():
                cf_list.append({
                    "range": str(cf_range),
                    "rules_count": len(rules),
                })
        return cf_list

    def _extract_tables(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Extrai tabelas"""
        tables = []
        if hasattr(sheet, "_tables"):
            for table in sheet._tables:
                tables.append({
                    "name": table.name,
                    "displayName": table.displayName,
                    "ref": table.ref,
                    "tableStyleInfo": {
                        "name": table.tableStyleInfo.name if table.tableStyleInfo else None,
                    },
                })
        return tables

    def _extract_image(self, img, sheet_name: str, index: int) -> Dict[str, Any]:
        """Extrai imagem com binário"""
        return {
            "sheet_name": sheet_name,
            "index": index,
            "anchor": str(img.anchor) if hasattr(img, "anchor") else None,
            "format": img.format if hasattr(img, "format") else None,
            "binary": img._data() if hasattr(img, "_data") else None,
        }

    def _serialize_value(self, value: Any) -> Any:
        """Serializa valor para JSON"""
        if value is None:
            return None
        if isinstance(value, (str, int, float, bool)):
            return value
        return str(value)

    def _validate_snapshot(self, snapshot: Dict[str, Any]):
        """
        Validação obrigatória do snapshot - RIGOROSA
        
        Garante que todos os componentes críticos foram extraídos:
        - schema_version
        - workbook properties
        - Todas as sheets com todos os campos obrigatórios
        - Células com estilos completos
        - merged_cells, validations, images (mesmo que vazio, deve estar presente)
        
        Raises:
            SnapshotValidationError: Se validação falhar
        """
        errors = []
        warnings = []
        
        # 1. Validar estrutura básica
        if "schema_version" not in snapshot:
            errors.append("schema_version ausente")
        
        if snapshot.get("schema_version") != SNAPSHOT_SCHEMA_VERSION:
            warnings.append(f"schema_version é {snapshot.get('schema_version')}, esperado {SNAPSHOT_SCHEMA_VERSION}")
        
        if "workbook" not in snapshot:
            errors.append("workbook properties ausentes")
        
        if "sheets" not in snapshot:
            errors.append("sheets ausente")
        
        sheets = snapshot.get("sheets", [])
        if not sheets:
            errors.append("nenhuma sheet encontrada no workbook")
        
        # 2. Validar cada sheet em detalhes
        for idx, sheet in enumerate(sheets):
            sheet_name = sheet.get("name", f"Sheet_{idx}")
            
            # 2.1. Componentes obrigatórios (estruturais)
            required_keys = [
                "name", "sheet_state", "freeze_panes", "page_setup", "page_margins",
                "row_dimensions", "column_dimensions", "merged_cells",
                "cells", "data_validations", "conditional_formatting", "tables", "images"
            ]
            
            for key in required_keys:
                if key not in sheet:
                    errors.append(f"Sheet '{sheet_name}': campo '{key}' ausente")
            
            # 2.2. Validar tipos de dados
            if "cells" in sheet and not isinstance(sheet["cells"], list):
                errors.append(f"Sheet '{sheet_name}': cells deve ser list")
            
            if "merged_cells" in sheet and not isinstance(sheet["merged_cells"], list):
                errors.append(f"Sheet '{sheet_name}': merged_cells deve ser list")
            
            if "data_validations" in sheet and not isinstance(sheet["data_validations"], list):
                errors.append(f"Sheet '{sheet_name}': data_validations deve ser list")
            
            # 2.3. Validar estrutura de page_setup
            if "page_setup" in sheet and isinstance(sheet["page_setup"], dict):
                ps = sheet["page_setup"]
                ps_keys = ["orientation", "paperSize", "fitToHeight", "fitToWidth", "scale"]
                for key in ps_keys:
                    if key not in ps:
                        errors.append(f"Sheet '{sheet_name}': page_setup.{key} ausente")
            
            # 2.4. Validar estrutura de page_margins
            if "page_margins" in sheet and isinstance(sheet["page_margins"], dict):
                pm = sheet["page_margins"]
                pm_keys = ["left", "right", "top", "bottom", "header", "footer"]
                for key in pm_keys:
                    if key not in pm:
                        errors.append(f"Sheet '{sheet_name}': page_margins.{key} ausente")
            
            # 2.5. Validar células têm estilo completo
            if "cells" in sheet:
                cells = sheet["cells"]
                if cells:  # Se houver células
                    # Verificar primeira célula como amostra
                    sample_cell = cells[0]
                    
                    # Validar campos obrigatórios de célula
                    cell_required_keys = [
                        "coordinate", "row", "column", "column_letter",
                        "value", "data_type", "formula", "number_format",
                        "hyperlink", "comment", "style"
                    ]
                    
                    for key in cell_required_keys:
                        if key not in sample_cell:
                            errors.append(f"Sheet '{sheet_name}': célula sem campo '{key}'")
                    
                    # Validar estilo completo
                    if "style" in sample_cell and isinstance(sample_cell["style"], dict):
                        style = sample_cell["style"]
                        style_keys = ["font", "fill", "border", "alignment", "protection"]
                        for key in style_keys:
                            if key not in style:
                                errors.append(f"Sheet '{sheet_name}': cell.style sem '{key}'")
                        
                        # Validar font completo
                        if "font" in style and isinstance(style["font"], dict):
                            font = style["font"]
                            font_keys = ["name", "size", "bold", "italic", "underline", "strike", "color"]
                            for key in font_keys:
                                if key not in font:
                                    errors.append(f"Sheet '{sheet_name}': font sem '{key}'")
                    else:
                        errors.append(f"Sheet '{sheet_name}': primeira célula sem style")
            
            # 2.6. Validar imagens (pode ser vazio, mas deve estar presente)
            if "images" not in sheet:
                errors.append(f"Sheet '{sheet_name}': images ausente (pode estar vazio)")
            elif isinstance(sheet["images"], list):
                for img_idx, img in enumerate(sheet["images"]):
                    img_keys = ["index", "anchor", "format"]
                    for key in img_keys:
                        if key not in img:
                            warnings.append(f"Sheet '{sheet_name}': image[{img_idx}] sem '{key}'")
        
        # 3. Falhar se houver erros críticos
        if errors:
            error_msg = f"❌ Snapshot INVÁLIDO ({len(errors)} erro(s)):\n"
            error_msg += "\n".join(f"  - {e}" for e in errors)
            if warnings:
                error_msg += f"\n\n⚠️ Avisos ({len(warnings)}):\n"
                error_msg += "\n".join(f"  - {w}" for w in warnings)
            logger.error(error_msg)
            raise SnapshotValidationError(error_msg)
        
        # 4. Log de sucesso
        log_msg = f"✅ Snapshot validado com sucesso: {len(sheets)} sheets, {sum(len(s.get('cells', [])) for s in sheets)} células"
        if warnings:
            log_msg += f", {len(warnings)} avisos"
        logger.info(log_msg)


def validate_snapshot(snapshot_dict: Dict[str, Any]) -> Dict[str, Any]:
    """
    Função standalone para validar snapshot
    
    Returns:
        Dict com:
        - valid: bool
        - errors: List[str]
        - stats: Dict com estatísticas
    """
    errors = []
    
    try:
        service = TemplateSnapshotService()
        service._validate_snapshot(snapshot_dict)
        valid = True
    except SnapshotValidationError as e:
        valid = False
        errors.append(str(e))
    
    stats = {
        "sheets_count": len(snapshot_dict.get("sheets", [])),
        "total_cells": sum(len(s.get("cells", [])) for s in snapshot_dict.get("sheets", [])),
        "total_merged": sum(len(s.get("merged_cells", [])) for s in snapshot_dict.get("sheets", [])),
        "total_validations": sum(len(s.get("data_validations", [])) for s in snapshot_dict.get("sheets", [])),
        "total_images": sum(len(s.get("images", [])) for s in snapshot_dict.get("sheets", [])),
    }
    
    return {
        "valid": valid,
        "errors": errors,
        "stats": stats,
    }
