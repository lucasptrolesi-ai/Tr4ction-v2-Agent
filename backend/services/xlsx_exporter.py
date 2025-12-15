"""
Serviço de exportação para Excel (XLSX)
Gera arquivos Excel a partir dos dados do banco
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, List, Any


def generate_xlsx(trail, steps: List, answers_by_step: Dict[str, Dict]) -> BytesIO:
    """
    Gera um arquivo XLSX com os dados da trilha.
    
    Args:
        trail: Objeto Trail com id e name
        steps: Lista de StepSchema ordenada
        answers_by_step: Dict {step_id: {campo: valor}}
    
    Returns:
        BytesIO stream com o arquivo XLSX
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove aba default
    
    # Estilos
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cria aba de resumo primeiro
    ws_summary = wb.create_sheet(title="Resumo")
    ws_summary["A1"] = f"Trilha: {trail.name if trail else 'N/A'}"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A3"] = "Etapas:"
    ws_summary["A3"].font = label_font
    
    row_summary = 4
    for step in steps:
        step_answers = answers_by_step.get(step.step_id, {})
        status = "✓ Preenchido" if step_answers else "○ Pendente"
        ws_summary[f"A{row_summary}"] = f"  • {step.step_name}: {status}"
        row_summary += 1
    
    ws_summary.column_dimensions['A'].width = 50
    
    # Cria uma aba para cada etapa
    for step in steps:
        # Nome da aba (máximo 31 caracteres)
        sheet_name = step.step_name[:31] if step.step_name else f"Step {step.step_id}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Header da etapa
        ws.merge_cells('A1:B1')
        ws["A1"] = f"Etapa: {step.step_name}"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal='center')
        
        # Descrição se existir
        if step.description:
            ws["A2"] = step.description
            ws["A2"].font = Font(italic=True, size=10)
            ws.merge_cells('A2:B2')
        
        row = 4
        
        # Headers das colunas
        ws["A3"] = "Campo"
        ws["B3"] = "Resposta"
        ws["A3"].font = label_font
        ws["B3"].font = label_font
        ws["A3"].border = thin_border
        ws["B3"].border = thin_border
        
        # Pega o schema de campos
        schema_fields = []
        if step.schema:
            schema_fields = step.schema.get("fields", [])
        
        # Respostas desta etapa
        step_answers = answers_by_step.get(step.step_id, {})
        
        for field in schema_fields:
            # Pega a chave do campo (pode ser "key", "name" ou "id")
            key = field.get("key") or field.get("name") or field.get("id", "")
            label = field.get("label", key)
            
            # Campo (coluna A)
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = label_font
            ws[f"A{row}"].border = thin_border
            ws[f"A{row}"].alignment = Alignment(vertical='top', wrap_text=True)
            
            # Valor (coluna B)
            value = step_answers.get(key, "")
            ws[f"B{row}"] = str(value) if value else ""
            ws[f"B{row}"].font = value_font
            ws[f"B{row}"].border = thin_border
            ws[f"B{row}"].alignment = Alignment(vertical='top', wrap_text=True)
            
            row += 1
        
        # Se não tem fields no schema, mas tem respostas, mostra as respostas
        if not schema_fields and step_answers:
            for key, value in step_answers.items():
                ws[f"A{row}"] = key
                ws[f"A{row}"].font = label_font
                ws[f"A{row}"].border = thin_border
                
                ws[f"B{row}"] = str(value) if value else ""
                ws[f"B{row}"].font = value_font
                ws[f"B{row}"].border = thin_border
                
                row += 1
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 60
    
    # Salva em memória
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return stream


def generate_simple_xlsx(data: List[Dict[str, Any]], sheet_name: str = "Dados") -> BytesIO:
    """
    Gera um XLSX simples a partir de uma lista de dicionários.
    Útil para exportações genéricas.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if not data:
        ws["A1"] = "Sem dados"
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream
    
    # Headers (chaves do primeiro item)
    headers = list(data[0].keys())
    header_font = Font(bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
    
    # Dados
    for row_num, row_data in enumerate(data, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col, value=row_data.get(header, ""))
    
    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return stream
