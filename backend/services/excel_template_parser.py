"""
Excel Template Parser
=====================
Converts Excel sheets to pixel-perfect JSON schemas for frontend rendering.
Handles exact cell position calculations, dimension conversions, and field extraction.

Features:
- Excel cell → pixel position conversion
- Row height and column width parsing
- JSON schema generation with field metadata
- Supports all Excel data types (string, number, date, boolean)
- Preserves cell formatting info (useful for UI hints)
"""

import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, field, asdict
from enum import Enum
import logging

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ============================================================================
# EXCEL → PIXEL CONVERSION CONSTANTS
# ============================================================================

# Excel default column width: 8.43 in default font (Calibri 11)
# 1 column unit = approximately 7 pixels at 96 DPI
EXCEL_COLUMN_UNIT_TO_PIXELS = 7.0

# Excel default row height: 15 points
# 1 row unit (point) = 1.33 pixels at 96 DPI
EXCEL_ROW_POINT_TO_PIXELS = 1.33

# Excel uses 20 twips per point (twips = 1/20 point)
TWIPS_PER_POINT = 20

# Margins: default padding in pixels around cell content
CELL_PADDING_X = 2  # left/right padding
CELL_PADDING_Y = 2  # top/bottom padding


class FieldType(str, Enum):
    """Supported field types for template forms."""
    TEXT = "text"                  # Single line text
    TEXTAREA = "textarea"          # Multi-line text
    NUMBER = "number"              # Numeric input
    DECIMAL = "decimal"            # Decimal number (float)
    DATE = "date"                  # Date picker
    BOOLEAN = "boolean"            # Checkbox
    CURRENCY = "currency"          # Currency (number with formatting)
    PERCENTAGE = "percentage"      # Percentage value
    ENUM = "enum"                  # Dropdown/select
    EMAIL = "email"                # Email validation
    PHONE = "phone"                # Phone number
    URL = "url"                    # URL validation
    JSON = "json"                  # JSON object (advanced)


@dataclass
class CellPosition:
    """Exact position of a cell in pixels."""
    top: float                      # Y-coordinate (from top)
    left: float                     # X-coordinate (from left)
    width: float                    # Cell width in pixels
    height: float                   # Cell height in pixels

    def to_dict(self) -> Dict[str, float]:
        return asdict(self)


@dataclass
class FieldMetadata:
    """Complete metadata for an editable template field."""
    key: str                       # Unique field identifier (e.g., "persona_name")
    cell: str                      # Excel cell address (e.g., "A1", "I16")
    type: FieldType               # Input type (text, number, date, etc.)
    label: Optional[str] = None   # Human-readable label
    placeholder: Optional[str] = None  # Placeholder text
    required: bool = False        # Whether field is mandatory
    validation_rules: Dict[str, Any] = field(default_factory=dict)
    # Format: {"min": 5, "max": 50, "pattern": "^[A-Z]", "enum": [...]}
    help_text: Optional[str] = None    # Contextual help for the user
    section: Optional[str] = None      # Section grouping (e.g., "Identity", "Market")
    
    # Position and size (calculated from Excel)
    position: Optional[CellPosition] = None
    
    # Original cell value (for comparison/default)
    original_value: Optional[str] = None


@dataclass
class TemplateSchema:
    """Complete template structure for frontend rendering."""
    template_key: str              # Unique template identifier (e.g., "persona_01")
    sheet_name: str                # Original Excel sheet name
    
    # Sheet dimensions
    sheet_width: float             # Total width in pixels
    sheet_height: float            # Total height in pixels
    
    # Field definitions
    fields: List[FieldMetadata] = field(default_factory=list)
    
    # Metadata
    title: Optional[str] = None
    description: Optional[str] = None
    version: str = "1.0"
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to JSON-serializable dictionary."""
        result = {
            "template_key": self.template_key,
            "sheet_name": self.sheet_name,
            "sheet_width": self.sheet_width,
            "sheet_height": self.sheet_height,
            "fields": [
                {
                    **asdict(field),
                    "type": field.type.value,
                    "position": field.position.to_dict() if field.position else None,
                    "validation_rules": field.validation_rules or {},
                }
                for field in self.fields
            ],
            "title": self.title,
            "description": self.description,
            "version": self.version,
        }
        return result


class ExcelTemplateParser:
    """
    Converts Excel sheets to pixel-perfect template schemas.
    
    Usage:
        parser = ExcelTemplateParser("Template Q1.xlsx")
        
        # Parse a single sheet with field definitions
        schema = parser.parse_sheet(
            sheet_name="Persona",
            fields={
                "persona_name": {"cell": "B2", "type": "text", "label": "Persona Name"},
                "company_size": {"cell": "B5", "type": "number", "label": "Company Size"},
            }
        )
        
        # Export as JSON
        with open("persona_schema.json", "w") as f:
            json.dump(schema.to_dict(), f, indent=2)
    """
    
    def __init__(self, excel_path: str | Path):
        """Initialize parser with Excel file path."""
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        
        self.workbook = openpyxl.load_workbook(str(self.excel_path))
        self._column_width_cache: Dict[str, float] = {}
        self._row_height_cache: Dict[int, float] = {}
    
    def _get_column_width_pixels(self, worksheet: Worksheet, col_letter: str) -> float:
        """
        Get column width in pixels.
        
        Excel stores column width as a measurement in character units.
        Default is 8.43 (width of 8 characters in Calibri 11 point).
        """
        if col_letter in self._column_width_cache:
            return self._column_width_cache[col_letter]
        
        col_width = worksheet.column_dimensions[col_letter].width
        
        # If not explicitly set, use Excel default
        if col_width is None:
            col_width = 8.43
        
        # Convert to pixels: multiply by 7 (approximately 7 pixels per column unit)
        pixels = col_width * EXCEL_COLUMN_UNIT_TO_PIXELS
        
        self._column_width_cache[col_letter] = pixels
        return pixels
    
    def _get_row_height_pixels(self, worksheet: Worksheet, row_num: int) -> float:
        """
        Get row height in pixels.
        
        Excel stores row height in points (1 point = 1/72 inch).
        Default is 15 points.
        At 96 DPI: 1 point ≈ 1.33 pixels
        """
        if row_num in self._row_height_cache:
            return self._row_height_cache[row_num]
        
        row_height = worksheet.row_dimensions[row_num].height
        
        # If not explicitly set, use Excel default
        if row_height is None:
            row_height = 15  # Default Excel row height in points
        
        # Convert to pixels: multiply by 1.33 (96 DPI conversion)
        pixels = row_height * EXCEL_ROW_POINT_TO_PIXELS
        
        self._row_height_cache[row_num] = pixels
        return pixels
    
    def get_cell_position(
        self,
        worksheet: Worksheet,
        cell_address: str
    ) -> CellPosition:
        """
        Calculate exact pixel position of a cell.
        
        Args:
            worksheet: openpyxl worksheet object
            cell_address: Cell address (e.g., "A1", "I16", "Z99")
        
        Returns:
            CellPosition with top, left, width, height in pixels
        """
        # Parse cell address
        col_index = column_index_from_string(cell_address.rstrip('0123456789'))
        row_num = int(cell_address.lstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))
        
        col_letter = get_column_letter(col_index)
        
        # Calculate cumulative left position
        left = 0.0
        for col in range(1, col_index):
            left += self._get_column_width_pixels(worksheet, get_column_letter(col))
        
        # Calculate cumulative top position
        top = 0.0
        for row in range(1, row_num):
            top += self._get_row_height_pixels(worksheet, row)
        
        # Get this cell's dimensions
        width = self._get_column_width_pixels(worksheet, col_letter)
        height = self._get_row_height_pixels(worksheet, row_num)
        
        logger.debug(f"Cell {cell_address}: top={top}, left={left}, width={width}, height={height}")
        
        return CellPosition(top=top, left=left, width=width, height=height)
    
    def _infer_field_type(self, cell) -> FieldType:
        """Infer field type from cell value and formatting."""
        if cell.value is None:
            return FieldType.TEXT
        
        # Check data type
        if isinstance(cell.value, bool):
            return FieldType.BOOLEAN
        elif isinstance(cell.value, (int, float)):
            # Check for percentage format
            if cell.number_format and '%' in cell.number_format:
                return FieldType.PERCENTAGE
            # Check for currency format
            elif cell.number_format and any(c in cell.number_format for c in '$€¥'):
                return FieldType.CURRENCY
            # Check for decimal format
            elif isinstance(cell.value, float):
                return FieldType.DECIMAL
            else:
                return FieldType.NUMBER
        elif isinstance(cell.value, str):
            # Check for email-like pattern (simple heuristic)
            if '@' in cell.value and '.' in cell.value:
                return FieldType.EMAIL
            # Check for URL pattern
            elif cell.value.startswith(('http://', 'https://', 'www.')):
                return FieldType.URL
            # Check for phone-like pattern (digits and common separators)
            elif all(c in '0123456789-() +' for c in cell.value):
                return FieldType.PHONE
            else:
                return FieldType.TEXT
        else:
            return FieldType.TEXT
    
    def parse_sheet(
        self,
        sheet_name: str,
        fields: Dict[str, Dict[str, Any]],
        template_key: Optional[str] = None,
        title: Optional[str] = None,
        description: Optional[str] = None,
    ) -> TemplateSchema:
        """
        Parse a single Excel sheet and generate template schema.
        
        Args:
            sheet_name: Name of worksheet to parse
            fields: Dict mapping field_key → field_config
                    Example: {
                        "persona_name": {
                            "cell": "B2",
                            "type": "text",
                            "label": "Persona Name",
                            "placeholder": "e.g., Young Urban Professional",
                            "required": True,
                            "help_text": "Who is this persona?"
                        }
                    }
            template_key: Unique identifier for this template (auto-generated if None)
            title: Human-readable title
            description: Description of this template
        
        Returns:
            TemplateSchema object with all field positions and metadata
        """
        # Get worksheet
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self.workbook.sheetnames}")
        
        worksheet = self.workbook[sheet_name]
        
        # Generate template_key if not provided
        if template_key is None:
            template_key = sheet_name.lower().replace(' ', '_')
        
        # Calculate sheet dimensions (use last row and column with content)
        max_row = worksheet.max_row or 1
        max_col = worksheet.max_column or 1
        
        sheet_width = 0.0
        for col in range(1, max_col + 1):
            sheet_width += self._get_column_width_pixels(worksheet, get_column_letter(col))
        
        sheet_height = 0.0
        for row in range(1, max_row + 1):
            sheet_height += self._get_row_height_pixels(worksheet, row)
        
        logger.info(f"Parsing sheet '{sheet_name}': {max_row} rows × {max_col} cols")
        logger.info(f"Sheet dimensions: {sheet_width}px × {sheet_height}px")
        
        # Parse field definitions
        field_list: List[FieldMetadata] = []
        
        for field_key, field_config in fields.items():
            cell_address = field_config["cell"]
            
            # Get cell object
            cell = worksheet[cell_address]
            
            # Get position
            position = self.get_cell_position(worksheet, cell_address)
            
            # Infer type if not specified
            field_type = field_config.get("type")
            if isinstance(field_type, str):
                try:
                    field_type = FieldType(field_type)
                except ValueError:
                    logger.warning(f"Invalid type '{field_type}' for field {field_key}, using TEXT")
                    field_type = FieldType.TEXT
            elif field_type is None:
                field_type = self._infer_field_type(cell)
            
            # Build validation rules
            validation_rules = field_config.get("validation_rules", {})
            
            # Build field metadata
            metadata = FieldMetadata(
                key=field_key,
                cell=cell_address,
                type=field_type,
                label=field_config.get("label", field_key),
                placeholder=field_config.get("placeholder"),
                required=field_config.get("required", False),
                validation_rules=validation_rules,
                help_text=field_config.get("help_text"),
                section=field_config.get("section"),
                position=position,
                original_value=str(cell.value) if cell.value is not None else None,
            )
            
            field_list.append(metadata)
            logger.debug(f"Field {field_key}: cell={cell_address}, type={field_type.value}")
        
        # Build template schema
        schema = TemplateSchema(
            template_key=template_key,
            sheet_name=sheet_name,
            sheet_width=sheet_width,
            sheet_height=sheet_height,
            fields=field_list,
            title=title or sheet_name,
            description=description,
        )
        
        logger.info(f"Generated schema for {len(field_list)} fields")
        
        return schema
    
    def export_schema_to_json(self, schema: TemplateSchema, output_path: str | Path) -> None:
        """Export template schema to JSON file."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(schema.to_dict(), f, indent=2, ensure_ascii=False)
        
        logger.info(f"Schema exported to {output_path}")
    
    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()


# ============================================================================
# EXAMPLE USAGE
# ============================================================================

def example_parse_persona_sheet():
    """Example: Parse Persona sheet from Template Q1.xlsx"""
    parser = ExcelTemplateParser("Template Q1.xlsx")
    
    # Define editable fields for Persona sheet
    persona_fields = {
        # Identity section
        "persona_name": {
            "cell": "B2",
            "type": "text",
            "label": "Persona Name",
            "placeholder": "e.g., 'Young Urban Professional'",
            "required": True,
            "section": "Identity",
            "help_text": "Give this persona a memorable name that captures their essence.",
        },
        "age_range": {
            "cell": "B3",
            "type": "text",
            "label": "Age Range",
            "placeholder": "e.g., '25-35'",
            "required": True,
            "section": "Identity",
        },
        "occupation": {
            "cell": "B4",
            "type": "text",
            "label": "Primary Occupation",
            "required": True,
            "section": "Identity",
        },
        
        # Psychographics section
        "values": {
            "cell": "B6",
            "type": "textarea",
            "label": "Core Values",
            "placeholder": "What matters most to them?",
            "section": "Psychographics",
            "help_text": "List 3-5 core values that drive decision-making.",
        },
        "pain_points": {
            "cell": "B7",
            "type": "textarea",
            "label": "Main Pain Points",
            "placeholder": "What challenges do they face daily?",
            "section": "Psychographics",
            "required": True,
        },
        "goals": {
            "cell": "B8",
            "type": "textarea",
            "label": "Primary Goals",
            "placeholder": "What do they want to achieve?",
            "section": "Psychographics",
            "required": True,
        },
        
        # Media & Communication
        "preferred_channels": {
            "cell": "B10",
            "type": "textarea",
            "label": "Preferred Communication Channels",
            "placeholder": "Social media, email, phone, in-person?",
            "section": "Communication",
        },
        "content_preferences": {
            "cell": "B11",
            "type": "text",
            "label": "Content Type Preferences",
            "placeholder": "Video, articles, podcasts, infographics?",
            "section": "Communication",
        },
    }
    
    # Parse the sheet
    schema = parser.parse_sheet(
        sheet_name="Persona",
        fields=persona_fields,
        template_key="persona_01",
        title="Customer Persona Template",
        description="Define a detailed customer persona for product/market fit analysis."
    )
    
    # Export to JSON
    parser.export_schema_to_json(schema, "persona_schema.json")
    parser.close()
    
    print(json.dumps(schema.to_dict(), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Run example
    try:
        example_parse_persona_sheet()
    except FileNotFoundError:
        print("Note: This example requires 'Template Q1.xlsx' in the current directory")
        print("\nExample usage:")
        print("""
from services.excel_template_parser import ExcelTemplateParser

parser = ExcelTemplateParser("path/to/Template Q1.xlsx")
schema = parser.parse_sheet(
    sheet_name="Persona",
    fields={
        "persona_name": {
            "cell": "B2",
            "type": "text",
            "label": "Persona Name"
        }
    }
)
        """)
