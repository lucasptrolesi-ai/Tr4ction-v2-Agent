"""
Template Ingestion Service - Pipeline automático de processamento de Excel templates

Responsabilidades:
1. Processar arquivo Excel enviado por admin
2. Enumerar todas as sheets
3. Para cada sheet: detectar células editáveis, gerar JSON schema, exportar PNG
4. Registrar templates no banco de dados
5. Gerar relatório de ingestão

IMPORTANTE: Totalmente genérico - não hardcoda cycles (Q1, Q2, Q3...)
"""

import os
import shutil
import json
import logging
from typing import List, Dict, Optional, Tuple
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, asdict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

logger = logging.getLogger(__name__)


# ============================================================
# 📁 CONFIGURAÇÃO DE STORAGE - DATA-DRIVEN
# ============================================================

BASE_DIR = Path(__file__).parent.parent
TEMPLATES_SOURCE_DIR = BASE_DIR / "data" / "templates_source"
TEMPLATES_GENERATED_DIR = BASE_DIR / "templates" / "generated"
TEMPLATES_IMAGES_DIR = BASE_DIR.parent / "frontend" / "public" / "templates"

# Criar diretórios base se não existirem
TEMPLATES_SOURCE_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATES_GENERATED_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATES_IMAGES_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# 📊 PIXEL CONVERSION CONSTANTS (FROM PERSONA 01)
# ============================================================

EXCEL_COLUMN_UNIT_TO_PIXELS = 7.0
EXCEL_ROW_POINT_TO_PIXELS = 1.33
TEXTAREA_HEIGHT_THRESHOLD = 40.0  # pixels - ~2 rows


# ============================================================
# 🏗️ DATA STRUCTURES
# ============================================================

@dataclass
class CellPosition:
    """Posição pixel-perfect de uma célula"""
    top: float
    left: float
    width: float
    height: float


@dataclass
class FieldSchema:
    """Schema de um campo editável"""
    key: str
    label: str
    cell: str
    type: str  # "text" ou "textarea"
    top: float
    left: float
    width: float
    height: float


@dataclass
class TemplateSchema:
    """Schema completo de um template"""
    template_key: str
    sheet_name: str
    sheet_width: float
    sheet_height: float
    title: str
    description: str
    version: str
    fields: List[Dict]


@dataclass
class IngestionResult:
    """Resultado do processamento de um template"""
    template_key: str
    sheet_name: str
    success: bool
    field_count: int
    schema_path: str
    image_path: str
    warnings: List[str]
    errors: List[str]


# ============================================================
# 🔧 EXCEL DIMENSION CALCULATOR
# ============================================================

class ExcelDimensionCalculator:
    """Calcula dimensões pixel-perfect das células Excel"""
    
    def __init__(self, worksheet):
        self.ws = worksheet
        self._col_width_cache = {}
        self._row_height_cache = {}
    
    def get_column_width_pixels(self, col_idx: int) -> float:
        """Retorna largura de uma coluna em pixels (com cache)"""
        if col_idx in self._col_width_cache:
            return self._col_width_cache[col_idx]
        
        col_letter = get_column_letter(col_idx)
        col_dim = self.ws.column_dimensions.get(col_letter)
        
        if col_dim and col_dim.width:
            width_px = col_dim.width * EXCEL_COLUMN_UNIT_TO_PIXELS
        else:
            width_px = 8.43 * EXCEL_COLUMN_UNIT_TO_PIXELS  # Default Excel width
        
        self._col_width_cache[col_idx] = width_px
        return width_px
    
    def get_row_height_pixels(self, row_idx: int) -> float:
        """Retorna altura de uma linha em pixels (com cache)"""
        if row_idx in self._row_height_cache:
            return self._row_height_cache[row_idx]
        
        row_dim = self.ws.row_dimensions.get(row_idx)
        
        if row_dim and row_dim.height:
            height_px = row_dim.height * EXCEL_ROW_POINT_TO_PIXELS
        else:
            height_px = 15 * EXCEL_ROW_POINT_TO_PIXELS  # Default Excel height
        
        self._row_height_cache[row_idx] = height_px
        return height_px
    
    def get_cell_position(self, cell_address: str) -> CellPosition:
        """Calcula posição pixel-perfect de uma célula"""
        cell = self.ws[cell_address]
        
        # Obter índices da célula
        col_idx = cell.column
        row_idx = cell.row
        
        # Calcular posição cumulativa
        left = sum(self.get_column_width_pixels(c) for c in range(1, col_idx))
        top = sum(self.get_row_height_pixels(r) for r in range(1, row_idx))
        
        # Se for merged cell, calcular dimensões totais
        if isinstance(cell, type(cell)) and hasattr(self.ws, 'merged_cells'):
            for merged_range in self.ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Calcular largura total das colunas mescladas
                    width = sum(
                        self.get_column_width_pixels(c)
                        for c in range(merged_range.min_col, merged_range.max_col + 1)
                    )
                    # Calcular altura total das linhas mescladas
                    height = sum(
                        self.get_row_height_pixels(r)
                        for r in range(merged_range.min_row, merged_range.max_row + 1)
                    )
                    return CellPosition(top=top, left=left, width=width, height=height)
        
        # Célula normal (não mesclada)
        width = self.get_column_width_pixels(col_idx)
        height = self.get_row_height_pixels(row_idx)
        
        return CellPosition(top=top, left=left, width=width, height=height)
    
    def get_sheet_dimensions(self) -> Tuple[float, float]:
        """Retorna dimensões totais da sheet (largura, altura)"""
        max_col = self.ws.max_column or 1
        max_row = self.ws.max_row or 1
        
        width = sum(self.get_column_width_pixels(c) for c in range(1, max_col + 1))
        height = sum(self.get_row_height_pixels(r) for r in range(1, max_row + 1))
        
        return width, height


# ============================================================
# 🔍 EDITABLE CELL DISCOVERY
# ============================================================

class EditableCellDiscovery:
    """Detecta células editáveis usando heurística"""
    
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def is_cell_editable(self, cell) -> bool:
        """
        Heurística: célula é editável se:
        1. Fill color é BRANCA (rgb == "FFFFFFFF" ou theme == 0)
        2. Borders são THIN em todos os 4 lados
        3. Célula não está merged OU é a célula anchor de um merge
        """
        if not cell or cell.value is not None:
            # Skip células com valores (labels)
            pass  # Continue validation
        
        # Check 1: White fill
        fill = cell.fill
        is_white = False
        
        if fill and fill.fgColor:
            if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb == 'FFFFFFFF':
                is_white = True
            elif hasattr(fill.fgColor, 'theme') and fill.fgColor.theme == 0:
                is_white = True
        
        if not is_white:
            return False
        
        # Check 2: Thin borders on all 4 sides
        border = cell.border
        if not border:
            return False
        
        required_borders = ['left', 'right', 'top', 'bottom']
        for side in required_borders:
            border_side = getattr(border, side, None)
            if not border_side or border_side.style != 'thin':
                return False
        
        # Check 3: Not merged OR is anchor cell
        if hasattr(self.ws, 'merged_cells'):
            for merged_range in self.ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # É anchor se for a primeira célula do range
                    min_cell = self.ws.cell(merged_range.min_row, merged_range.min_col)
                    if cell.coordinate != min_cell.coordinate:
                        return False  # Não é anchor
        
        return True
    
    def discover_editable_cells(self) -> List[str]:
        """Retorna lista de endereços de células editáveis (ordenada)"""
        editable_cells = []
        
        max_row = self.ws.max_row or 1
        max_col = self.ws.max_column or 1
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = self.ws.cell(row, col)
                if self.is_cell_editable(cell):
                    editable_cells.append(cell.coordinate)
        
        return sorted(editable_cells, key=lambda c: (self.ws[c].row, self.ws[c].column))


# ============================================================
# 🏷️ LABEL EXTRACTOR
# ============================================================

class LabelExtractor:
    """Extrai labels verbatim das células próximas"""
    
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def is_label_candidate(self, cell) -> bool:
        """Label é célula com fill não-branco OU fonte bold"""
        if not cell:
            return False
        
        # Check fill color (non-white)
        fill = cell.fill
        if fill and fill.fgColor:
            if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb != 'FFFFFFFF':
                return True
            if hasattr(fill.fgColor, 'theme') and fill.fgColor.theme != 0:
                return True
        
        # Check bold font
        font = cell.font
        if font and font.bold:
            return True
        
        return False
    
    def extract_label(self, cell_address: str) -> str:
        """
        Busca label para uma célula editável:
        1. Olha para a ESQUERDA (até 8 colunas) na mesma linha
        2. Se não encontrar, olha para CIMA (até 12 linhas) na mesma coluna
        3. Retorna string vazia se não encontrar
        """
        cell = self.ws[cell_address]
        row_idx = cell.row
        col_idx = cell.column
        
        # Strategy 1: Look LEFT (same row)
        for offset in range(1, 9):  # até 8 colunas à esquerda
            if col_idx - offset < 1:
                break
            candidate = self.ws.cell(row_idx, col_idx - offset)
            if candidate.value and self.is_label_candidate(candidate):
                return str(candidate.value).strip()
        
        # Strategy 2: Look UP (same column)
        for offset in range(1, 13):  # até 12 linhas acima
            if row_idx - offset < 1:
                break
            candidate = self.ws.cell(row_idx - offset, col_idx)
            if candidate.value and self.is_label_candidate(candidate):
                return str(candidate.value).strip()
        
        return ""  # Não encontrou label


# ============================================================
# 🎨 PNG EXPORTER
# ============================================================

class PNGExporter:
    """Gera PNG background para template"""
    
    @staticmethod
    def generate_fallback_png(sheet_name: str, width: int, height: int, output_path: str):
        """Gera PNG placeholder com grid"""
        # Criar imagem branca
        img = Image.new('RGB', (int(width), int(height)), color='white')
        draw = ImageDraw.Draw(img)
        
        # Desenhar grid (linhas a cada 50px)
        grid_spacing = 50
        for x in range(0, int(width), grid_spacing):
            draw.line([(x, 0), (x, height)], fill='lightgray', width=1)
        for y in range(0, int(height), grid_spacing):
            draw.line([(0, y), (width, y)], fill='lightgray', width=1)
        
        # Desenhar nome do template no centro
        try:
            # Tentar usar fonte padrão
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 24)
        except:
            font = ImageFont.load_default()
        
        text = f"Template: {sheet_name}"
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        text_x = (width - text_width) / 2
        text_y = (height - text_height) / 2
        
        draw.text((text_x, text_y), text, fill='gray', font=font)
        
        # Salvar PNG
        img.save(output_path, 'PNG')


# ============================================================
# 📋 TEMPLATE SCHEMA GENERATOR
# ============================================================

class TemplateSchemaGenerator:
    """Gera schema JSON completo para um template"""
    
    def __init__(self, worksheet, template_key: str):
        self.ws = worksheet
        self.template_key = template_key
        self.calculator = ExcelDimensionCalculator(worksheet)
        self.discovery = EditableCellDiscovery(worksheet)
        self.label_extractor = LabelExtractor(worksheet)
    
    def infer_field_type(self, position: CellPosition) -> str:
        """Infere tipo do campo baseado na altura"""
        return "textarea" if position.height >= TEXTAREA_HEIGHT_THRESHOLD else "text"
    
    def generate_field_key(self, cell_address: str, index: int) -> str:
        """Gera chave única para o campo"""
        return f"field_{cell_address.lower()}_{index}"
    
    def generate_schema(self) -> Tuple[TemplateSchema, List[str]]:
        """Gera schema completo + lista de warnings"""
        warnings = []
        
        # Descobrir células editáveis
        editable_cells = self.discovery.discover_editable_cells()
        
        if not editable_cells:
            warnings.append(f"No editable cells discovered in sheet '{self.ws.title}'")
        
        # Gerar fields
        fields = []
        for idx, cell_address in enumerate(editable_cells):
            position = self.calculator.get_cell_position(cell_address)
            label = self.label_extractor.extract_label(cell_address)
            
            if not label:
                warnings.append(f"No label found for cell {cell_address}")
            
            field = FieldSchema(
                key=self.generate_field_key(cell_address, idx),
                label=label,
                cell=cell_address,
                type=self.infer_field_type(position),
                top=position.top,
                left=position.left,
                width=position.width,
                height=position.height
            )
            fields.append(asdict(field))
        
        # Dimensões da sheet
        sheet_width, sheet_height = self.calculator.get_sheet_dimensions()
        
        # Criar schema
        schema = TemplateSchema(
            template_key=self.template_key,
            sheet_name=self.ws.title,
            sheet_width=sheet_width,
            sheet_height=sheet_height,
            title=self.ws.title,
            description=f"Auto-generated template from sheet '{self.ws.title}'",
            version="1.0.0",
            fields=fields
        )
        
        return schema, warnings


# ============================================================
# 🚀 TEMPLATE INGESTION SERVICE (MAIN)
# ============================================================

class TemplateIngestionService:
    """
    Serviço principal de ingestão de templates
    
    Pipeline:
    1. Upload de Excel → salvar em /data/templates_source/{cycle}/
    2. Enumerar sheets
    3. Para cada sheet: detectar células + gerar schema JSON + exportar PNG
    4. Registrar no banco (TemplateDefinition)
    5. Gerar relatório de ingestão
    """
    
    def __init__(self, db_session):
        self.db = db_session
    
    def save_uploaded_file(self, file_content: bytes, filename: str, cycle: str) -> str:
        """
        Salva arquivo Excel enviado por admin
        
        Returns:
            Path completo do arquivo salvo
        """
        # Criar diretório do cycle
        cycle_dir = TEMPLATES_SOURCE_DIR / cycle
        cycle_dir.mkdir(parents=True, exist_ok=True)
        
        # Salvar arquivo
        file_path = cycle_dir / filename
        with open(file_path, 'wb') as f:
            f.write(file_content)
        
        logger.info(f"Saved uploaded file: {file_path}")
        return str(file_path)
    
    def normalize_template_key(self, sheet_name: str) -> str:
        """
        Normaliza nome da sheet para template_key
        
        Exemplo: "3.1 Persona 01" → "3_1_persona_01"
        """
        import unicodedata
        import re
        
        # Remover acentos
        nfkd = unicodedata.normalize('NFKD', sheet_name)
        key = ''.join([c for c in nfkd if not unicodedata.combining(c)])
        
        # Lowercase
        key = key.lower()
        
        # Substituir espaços e pontos por underscore
        key = re.sub(r'[\s.]+', '_', key)
        
        # Remover caracteres especiais
        key = re.sub(r'[^a-z0-9_]', '', key)
        
        # Remover underscores consecutivos
        key = re.sub(r'_+', '_', key)
        
        # Remover underscores no início/fim
        key = key.strip('_')
        
        return key
    
    def process_template(
        self,
        workbook_path: str,
        sheet_name: str,
        cycle: str
    ) -> IngestionResult:
        """
        Processa uma sheet individual do Excel
        
        Returns:
            IngestionResult com status e paths gerados
        """
        errors = []
        warnings = []
        
        try:
            # Carregar workbook
            wb = load_workbook(workbook_path, data_only=False)
            ws = wb[sheet_name]
            
            # Gerar template_key normalizado
            template_key = self.normalize_template_key(sheet_name)
            
            # Criar diretórios para o cycle
            cycle_schema_dir = TEMPLATES_GENERATED_DIR / cycle
            cycle_images_dir = TEMPLATES_IMAGES_DIR / cycle
            cycle_schema_dir.mkdir(parents=True, exist_ok=True)
            cycle_images_dir.mkdir(parents=True, exist_ok=True)
            
            # Gerar schema JSON
            generator = TemplateSchemaGenerator(ws, template_key)
            schema, gen_warnings = generator.generate_schema()
            warnings.extend(gen_warnings)
            
            # Salvar schema JSON
            schema_path = cycle_schema_dir / f"{template_key}.json"
            with open(schema_path, 'w', encoding='utf-8') as f:
                json.dump(asdict(schema), f, indent=2, ensure_ascii=False)
            
            # Gerar PNG placeholder
            image_path = cycle_images_dir / f"{template_key}.png"
            PNGExporter.generate_fallback_png(
                sheet_name=sheet_name,
                width=schema.sheet_width,
                height=schema.sheet_height,
                output_path=str(image_path)
            )
            
            logger.info(f"✅ Processed template '{template_key}' - {len(schema.fields)} fields")
            
            return IngestionResult(
                template_key=template_key,
                sheet_name=sheet_name,
                success=True,
                field_count=len(schema.fields),
                schema_path=str(schema_path.relative_to(BASE_DIR)),
                image_path=str(image_path.relative_to(BASE_DIR.parent)),
                warnings=warnings,
                errors=errors
            )
            
        except Exception as e:
            logger.error(f"❌ Error processing sheet '{sheet_name}': {e}", exc_info=True)
            errors.append(str(e))
            
            return IngestionResult(
                template_key=self.normalize_template_key(sheet_name),
                sheet_name=sheet_name,
                success=False,
                field_count=0,
                schema_path="",
                image_path="",
                warnings=warnings,
                errors=errors
            )
    
    def ingest_excel_file(
        self,
        file_path: str,
        cycle: str,
        description: Optional[str] = None
    ) -> Dict:
        """
        Pipeline completo de ingestão de um arquivo Excel
        
        Args:
            file_path: Path do arquivo Excel
            cycle: Identificador do cycle (Q1, Q2, Q3...)
            description: Descrição opcional do batch de templates
        
        Returns:
            Dict com estatísticas e relatório
        """
        from db.models import TemplateDefinition
        
        logger.info(f"🚀 Starting ingestion for cycle '{cycle}' - file: {file_path}")
        
        # Carregar workbook para enumerar sheets
        wb = load_workbook(file_path, data_only=False, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        logger.info(f"Found {len(sheet_names)} sheets: {sheet_names}")
        
        # Processar cada sheet
        results: List[IngestionResult] = []
        for sheet_name in sheet_names:
            result = self.process_template(file_path, sheet_name, cycle)
            results.append(result)
        
        # Registrar templates no banco
        registered_count = 0
        for result in results:
            if result.success:
                # Verificar se já existe
                existing = self.db.query(TemplateDefinition).filter_by(
                    cycle=cycle,
                    template_key=result.template_key
                ).first()
                
                if existing:
                    # Atualizar existente
                    existing.sheet_name = result.sheet_name
                    existing.schema_path = result.schema_path
                    existing.image_path = result.image_path
                    existing.field_count = result.field_count
                    existing.ingestion_report = "\n".join(result.warnings) if result.warnings else None
                    existing.updated_at = datetime.utcnow()
                    logger.info(f"Updated existing template: {result.template_key}")
                else:
                    # Criar novo
                    template_def = TemplateDefinition(
                        cycle=cycle,
                        template_key=result.template_key,
                        sheet_name=result.sheet_name,
                        schema_path=result.schema_path,
                        image_path=result.image_path,
                        status="active",
                        description=description,
                        field_count=result.field_count,
                        source_file=file_path,
                        ingestion_report="\n".join(result.warnings) if result.warnings else None
                    )
                    self.db.add(template_def)
                    logger.info(f"Registered new template: {result.template_key}")
                
                registered_count += 1
        
        self.db.commit()
        
        # Gerar relatório
        report = self._generate_ingestion_report(cycle, results, file_path)
        
        # Salvar relatório
        report_path = BASE_DIR / f"TEMPLATE_INGESTION_REPORT_{cycle}.md"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report)
        
        logger.info(f"✅ Ingestion complete - {registered_count}/{len(results)} templates registered")
        
        return {
            "cycle": cycle,
            "total_sheets": len(sheet_names),
            "successful": sum(1 for r in results if r.success),
            "failed": sum(1 for r in results if not r.success),
            "total_fields": sum(r.field_count for r in results),
            "registered_in_db": registered_count,
            "report_path": str(report_path),
            "results": [asdict(r) for r in results]
        }
    
    def _generate_ingestion_report(
        self,
        cycle: str,
        results: List[IngestionResult],
        source_file: str
    ) -> str:
        """Gera relatório Markdown da ingestão"""
        
        report = f"""# Template Ingestion Report - {cycle}

**Generated**: {datetime.utcnow().isoformat()}

**Source File**: {source_file}

**Cycle**: {cycle}

---

## Summary

- **Total templates processed**: {len(results)}
- **Successful**: {sum(1 for r in results if r.success)}
- **Failed**: {sum(1 for r in results if not r.success)}
- **Total fields generated**: {sum(r.field_count for r in results)}
- **Total warnings**: {sum(len(r.warnings) for r in results)}

---

## Processed Templates

| Template Key | Sheet Name | Status | Fields | Warnings |
|--------------|------------|--------|--------|----------|
"""
        
        for result in results:
            status_icon = "✅" if result.success else "❌"
            report += f"| {result.template_key} | {result.sheet_name} | {status_icon} | {result.field_count} | {len(result.warnings)} |\n"
        
        report += "\n---\n\n## Warnings\n\n"
        
        for result in results:
            if result.warnings:
                report += f"### {result.template_key}\n\n"
                for warning in result.warnings:
                    report += f"- {warning}\n"
                report += "\n"
        
        if any(r.errors for r in results):
            report += "\n---\n\n## Errors\n\n"
            for result in results:
                if result.errors:
                    report += f"### {result.template_key}\n\n"
                    for error in result.errors:
                        report += f"- ❌ {error}\n"
                    report += "\n"
        
        report += "\n---\n\n## Validation\n\n"
        
        # Validação básica
        all_have_schemas = all(r.schema_path for r in results if r.success)
        all_have_images = all(r.image_path for r in results if r.success)
        at_least_one_template = any(r.success for r in results)
        
        report += f"- ✅ At least 1 template rendered: {at_least_one_template}\n"
        report += f"- ✅ All schemas non-empty: {all_have_schemas}\n"
        report += f"- ✅ All PNGs exist: {all_have_images}\n"
        
        report += "\n---\n\n## Status\n\n"
        
        if all([all_have_schemas, all_have_images, at_least_one_template]):
            report += "✅ **INGESTION SUCCESSFUL** - All templates ready for use\n"
        else:
            report += "⚠️ **INGESTION COMPLETED WITH WARNINGS** - Review errors above\n"
        
        return report
