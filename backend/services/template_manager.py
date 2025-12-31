"""
Template Management Service
===========================
Handles loading, saving, and exporting template data.
Provides persistence layer for founder template responses.

Features:
- Load template schemas from JSON
- Persist user input per startup + template
- Export filled templates back to Excel
- Validation against schema
- History/versioning support
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
from functools import lru_cache

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

from services.excel_template_parser import TemplateSchema, FieldMetadata

logger = logging.getLogger(__name__)


class TemplateDataService:
    """
    Service for managing template data persistence and export.
    
    Data is stored in:
    - JSON files: `data/templates/{startup_id}/{template_key}/{version}.json`
    
    Structure:
    {
        "template_key": "persona_01",
        "startup_id": "startup_uuid",
        "data": {
            "persona_name": "Young Urban Professional",
            "age_range": "25-35",
            ...
        },
        "created_at": "2024-01-15T10:30:00",
        "updated_at": "2024-01-15T10:30:00",
        "version": 1
    }
    """
    
    def __init__(self, data_dir: str | Path = "data/templates"):
        """Initialize service with data directory."""
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        self.schemas_dir = Path("data/schemas")
        self.schemas_dir.mkdir(parents=True, exist_ok=True)
    
    @lru_cache(maxsize=128)
    def load_schema(self, template_key: str) -> TemplateSchema:
        """
        Load template schema from JSON file.
        Results are cached for performance.
        """
        schema_path = self.schemas_dir / f"{template_key}.json"
        
        if not schema_path.exists():
            raise FileNotFoundError(f"Schema not found: {schema_path}")
        
        with open(schema_path, 'r', encoding='utf-8') as f:
            schema_dict = json.load(f)
        
        # Reconstruct TemplateSchema from dict
        schema = TemplateSchema(
            template_key=schema_dict["template_key"],
            sheet_name=schema_dict["sheet_name"],
            sheet_width=schema_dict["sheet_width"],
            sheet_height=schema_dict["sheet_height"],
            fields=[
                FieldMetadata(
                    key=field["key"],
                    cell=field["cell"],
                    type=field["type"],
                    label=field.get("label"),
                    placeholder=field.get("placeholder"),
                    required=field.get("required", False),
                    validation_rules=field.get("validation_rules", {}),
                    help_text=field.get("help_text"),
                    section=field.get("section"),
                    original_value=field.get("original_value"),
                )
                for field in schema_dict.get("fields", [])
            ],
            title=schema_dict.get("title"),
            description=schema_dict.get("description"),
            version=schema_dict.get("version", "1.0"),
        )
        
        logger.info(f"Loaded schema: {template_key}")
        return schema
    
    def save_schema(self, schema: TemplateSchema) -> None:
        """Save template schema to JSON file."""
        schema_path = self.schemas_dir / f"{schema.template_key}.json"
        
        with open(schema_path, 'w', encoding='utf-8') as f:
            json.dump(schema.to_dict(), f, indent=2, ensure_ascii=False)
        
        # Clear cache
        self.load_schema.cache_clear()
        
        logger.info(f"Saved schema: {schema.template_key}")
    
    def _get_data_path(self, startup_id: str, template_key: str, version: int = 1) -> Path:
        """Get path for template data file."""
        path = self.data_dir / startup_id / template_key / f"v{version}.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        return path
    
    def load_template_data(
        self,
        startup_id: str,
        template_key: str,
        version: int = 1
    ) -> Optional[Dict[str, Any]]:
        """
        Load saved template data for a startup.
        
        Returns:
            Dict with structure {template_key, startup_id, data, created_at, updated_at, version}
            or None if no data exists
        """
        data_path = self._get_data_path(startup_id, template_key, version)
        
        if not data_path.exists():
            logger.info(f"No saved data for {startup_id}/{template_key}/v{version}")
            return None
        
        with open(data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        logger.info(f"Loaded template data: {startup_id}/{template_key}")
        return data
    
    def save_template_data(
        self,
        startup_id: str,
        template_key: str,
        data: Dict[str, Any],
        auto_version: bool = True
    ) -> Dict[str, Any]:
        """
        Save template data for a startup.
        
        Args:
            startup_id: Unique startup identifier
            template_key: Template identifier
            data: Form data (field_key → value)
            auto_version: If True, auto-increment version, else use v1
        
        Returns:
            Saved data object with metadata
        """
        # Load existing to determine next version
        if auto_version:
            existing = self.load_template_data(startup_id, template_key)
            version = (existing.get("version", 0) + 1) if existing else 1
        else:
            version = 1
        
        # Create data object
        now = datetime.utcnow().isoformat()
        saved_data = {
            "template_key": template_key,
            "startup_id": startup_id,
            "data": data,
            "created_at": existing.get("created_at", now) if not auto_version else now,
            "updated_at": now,
            "version": version,
        }
        
        # Write to file
        data_path = self._get_data_path(startup_id, template_key, version)
        with open(data_path, 'w', encoding='utf-8') as f:
            json.dump(saved_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Saved template data v{version}: {startup_id}/{template_key}")
        return saved_data
    
    def list_template_versions(
        self,
        startup_id: str,
        template_key: str
    ) -> List[Dict[str, Any]]:
        """List all versions of a template for a startup."""
        template_dir = self.data_dir / startup_id / template_key
        
        if not template_dir.exists():
            return []
        
        versions = []
        for version_file in sorted(template_dir.glob("v*.json")):
            with open(version_file, 'r', encoding='utf-8') as f:
                versions.append(json.load(f))
        
        return versions
    
    def validate_data(
        self,
        template_key: str,
        data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Validate form data against template schema.
        
        Returns:
            {
                "valid": bool,
                "errors": [{"field": "...", "message": "..."}],
                "warnings": [{"field": "...", "message": "..."}]
            }
        """
        schema = self.load_schema(template_key)
        errors = []
        warnings = []
        
        # Check for required fields
        for field in schema.fields:
            if field.required and field.key not in data:
                errors.append({
                    "field": field.key,
                    "message": f"{field.label} is required"
                })
            
            # Validate against rules
            if field.key in data:
                value = data[field.key]
                
                # Check length constraints
                if field.validation_rules.get("min") and len(str(value)) < field.validation_rules["min"]:
                    errors.append({
                        "field": field.key,
                        "message": f"Minimum length is {field.validation_rules['min']}"
                    })
                
                if field.validation_rules.get("max") and len(str(value)) > field.validation_rules["max"]:
                    errors.append({
                        "field": field.key,
                        "message": f"Maximum length is {field.validation_rules['max']}"
                    })
                
                # Check enum values
                if field.validation_rules.get("enum"):
                    if value not in field.validation_rules["enum"]:
                        errors.append({
                            "field": field.key,
                            "message": f"Invalid value. Must be one of: {field.validation_rules['enum']}"
                        })
                
                # Check pattern (regex)
                if field.validation_rules.get("pattern"):
                    import re
                    pattern = field.validation_rules["pattern"]
                    if not re.match(pattern, str(value)):
                        errors.append({
                            "field": field.key,
                            "message": f"Invalid format"
                        })
        
        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
        }
    
    def export_to_excel(
        self,
        startup_id: str,
        template_key: str,
        original_excel_path: str | Path,
        output_excel_path: str | Path,
        version: int = 1
    ) -> Path:
        """
        Export template data back to Excel file.
        Writes values into the exact same cells they came from.
        
        Args:
            startup_id: Startup identifier
            template_key: Template identifier
            original_excel_path: Path to original Excel file
            output_excel_path: Where to save the filled Excel
            version: Data version to export
        
        Returns:
            Path to exported Excel file
        """
        # Load schema and data
        schema = self.load_schema(template_key)
        data = self.load_template_data(startup_id, template_key, version)
        
        if not data:
            raise ValueError(f"No data found for {startup_id}/{template_key}/v{version}")
        
        # Load original Excel
        workbook = openpyxl.load_workbook(str(original_excel_path))
        worksheet = workbook[schema.sheet_name]
        
        # Write field values into cells
        for field in schema.fields:
            if field.key in data["data"]:
                value = data["data"][field.key]
                cell = worksheet[field.cell]
                cell.value = value
                
                # Add light background to indicate filled cells (optional UX hint)
                cell.fill = PatternFill(
                    start_color="FFFACD",  # Light yellow
                    end_color="FFFACD",
                    fill_type="solid"
                )
                
                logger.debug(f"Wrote {field.key} = {value} to cell {field.cell}")
        
        # Add metadata sheet
        if "Metadata" in workbook.sheetnames:
            del workbook["Metadata"]
        
        metadata_sheet = workbook.create_sheet("Metadata", 0)
        metadata_sheet["A1"] = "Export Info"
        metadata_sheet["A2"] = f"Template: {template_key}"
        metadata_sheet["A3"] = f"Startup: {startup_id}"
        metadata_sheet["A4"] = f"Version: {version}"
        metadata_sheet["A5"] = f"Exported: {datetime.utcnow().isoformat()}"
        
        # Save
        output_path = Path(output_excel_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(output_path))
        workbook.close()
        
        logger.info(f"Exported to Excel: {output_path}")
        return output_path


class TemplateManager:
    """High-level interface for template operations."""
    
    def __init__(self, data_service: Optional[TemplateDataService] = None):
        self.data_service = data_service or TemplateDataService()
    
    def get_template_for_founder(
        self,
        startup_id: str,
        template_key: str
    ) -> Dict[str, Any]:
        """
        Get template schema + saved data (if exists) for founder.
        
        Returns:
            {
                "schema": {...},
                "saved_data": {...} or null,
                "versions": [...]
            }
        """
        schema = self.data_service.load_schema(template_key)
        saved_data = self.data_service.load_template_data(startup_id, template_key)
        versions = self.data_service.list_template_versions(startup_id, template_key)
        
        return {
            "schema": schema.to_dict(),
            "saved_data": saved_data,
            "versions": versions,
        }
    
    def save_founder_response(
        self,
        startup_id: str,
        template_key: str,
        data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Save founder's template response with validation."""
        # Validate
        validation = self.data_service.validate_data(template_key, data)
        if not validation["valid"]:
            raise ValueError(f"Validation failed: {validation['errors']}")
        
        # Save
        saved = self.data_service.save_template_data(startup_id, template_key, data)
        return saved
    
    def export_founder_template(
        self,
        startup_id: str,
        template_key: str,
        original_excel_path: str | Path,
        output_dir: str | Path = "exports"
    ) -> Path:
        """Export founder's template response to Excel."""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        output_path = output_dir / f"{startup_id}_{template_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return self.data_service.export_to_excel(
            startup_id=startup_id,
            template_key=template_key,
            original_excel_path=original_excel_path,
            output_excel_path=output_path
        )
