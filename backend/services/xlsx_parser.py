"""
Parser de Template Excel para Schema JSON
Converte arquivos XLSX em schemas de formulário automaticamente
"""
from openpyxl import load_workbook
from io import BytesIO
from typing import List, Dict, Any
import re


def sanitize_key(text: str) -> str:
    """
    Converte texto em chave válida para JSON/formulário
    Ex: "Campo Principal" -> "campo_principal"
    """
    if not text:
        return ""
    
    # Remove caracteres especiais, mantém letras, números e espaços
    clean = re.sub(r'[^\w\s]', '', str(text))
    # Converte para minúsculas e substitui espaços por underscore
    return clean.lower().strip().replace(' ', '_')


def detect_field_type(label: str, value: Any = None) -> str:
    """
    Detecta o tipo de campo baseado no label ou valor
    """
    label_lower = label.lower() if label else ""
    
    # Detecta tipos específicos
    if any(word in label_lower for word in ['email', 'e-mail']):
        return "email"
    if any(word in label_lower for word in ['telefone', 'celular', 'phone']):
        return "tel"
    if any(word in label_lower for word in ['data', 'date', 'quando']):
        return "date"
    if any(word in label_lower for word in ['valor', 'preço', 'price', 'custo', 'investimento']):
        return "number"
    if any(word in label_lower for word in ['url', 'link', 'site', 'website']):
        return "url"
    if any(word in label_lower for word in ['descrição', 'description', 'detalhe', 'explique', 'descreva', 'observação']):
        return "textarea"
    
    # Default: textarea para campos longos, text para curtos
    if value and isinstance(value, str) and len(value) > 100:
        return "textarea"
    
    return "textarea"  # Default seguro para formulários de negócio


def parse_template_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parseia um arquivo Excel e extrai os schemas de cada aba.
    
    Estrutura esperada do Excel:
    - Cada aba = uma etapa (step)
    - Coluna A = Label do campo
    - Coluna B = Valor de exemplo ou placeholder
    
    Args:
        file_bytes: Conteúdo do arquivo XLSX em bytes
    
    Returns:
        Lista de schemas para cada etapa
    """
    # Carrega workbook do bytes
    stream = BytesIO(file_bytes)
    wb = load_workbook(stream, data_only=True)
    
    steps = []
    order = 1
    
    for sheet in wb.worksheets:
        sheet_name = sheet.title.strip()
        
        # Ignora abas de configuração/metadados
        if sheet_name.lower() in ['config', 'metadata', 'instructions', 'instruções', 'readme']:
            continue
        
        fields = []
        seen_keys = set()  # Evita duplicatas
        
        # Percorre as linhas (assume A=label, B=valor/exemplo)
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            label_cell = row[0]
            value_cell = row[1] if len(row) > 1 else None
            
            # Ignora linhas vazias ou com células vazias na coluna A
            if not label_cell:
                continue
            
            label = str(label_cell).strip()
            
            # Ignora headers genéricos
            if label.lower() in ['campo', 'field', 'label', 'valor', 'value', 'resposta', 'answer']:
                continue
            
            # Gera chave única
            key = sanitize_key(label)
            if not key:
                continue
            
            # Se já existe, adiciona sufixo
            original_key = key
            counter = 1
            while key in seen_keys:
                key = f"{original_key}_{counter}"
                counter += 1
            seen_keys.add(key)
            
            # Detecta tipo do campo
            field_type = detect_field_type(label, value_cell)
            
            # Monta o campo
            field = {
                "name": key,  # Compatibilidade com frontend
                "key": key,
                "label": label,
                "type": field_type,
                "required": False,
                "placeholder": str(value_cell) if value_cell else f"Digite {label.lower()}..."
            }
            
            fields.append(field)
        
        # Só adiciona se tiver campos
        if fields:
            step_id = sanitize_key(sheet_name)
            
            steps.append({
                "step_id": step_id,
                "step_name": sheet_name,
                "order": order,
                "schema": {
                    "fields": fields
                }
            })
            order += 1
    
    return steps


def parse_single_sheet_xlsx(file_bytes: bytes, step_id: str, step_name: str) -> Dict[str, Any]:
    """
    Parseia um Excel de uma única aba/etapa.
    Útil quando admin quer fazer upload de uma etapa específica.
    """
    stream = BytesIO(file_bytes)
    wb = load_workbook(stream, data_only=True)
    
    # Pega a primeira aba
    sheet = wb.active
    
    fields = []
    seen_keys = set()
    
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        label_cell = row[0]
        value_cell = row[1] if len(row) > 1 else None
        
        if not label_cell:
            continue
        
        label = str(label_cell).strip()
        key = sanitize_key(label)
        
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)
        
        field_type = detect_field_type(label, value_cell)
        
        fields.append({
            "key": key,
            "label": label,
            "type": field_type,
            "required": False,
            "placeholder": str(value_cell) if value_cell else ""
        })
    
    return {
        "step_id": step_id,
        "step_name": step_name,
        "schema": {
            "fields": fields
        }
    }
