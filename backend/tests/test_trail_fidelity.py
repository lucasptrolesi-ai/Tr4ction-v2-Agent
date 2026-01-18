"""
Testes de Fidelidade de Trilha - Garantir que ordem é preservada

Testes:
1. Ordem das abas no Excel == ordem no sistema
2. Ordem das perguntas no Excel == ordem no sistema
3. Nenhuma pergunta é perdida
4. field_id é estável
5. Trilha é reprodutível
"""

import pytest
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.services.template_snapshot import TemplateSnapshotService
from app.services.question_extractor import QuestionExtractor
from app.services.trail_ingestion_service import TrailIngestionService, TrailIngestionError


# ============================================================
# FIXTURES
# ============================================================

@pytest.fixture
def trail_workbook_bytes():
    """Cria um workbook que simula uma trilha educacional"""
    
    wb = Workbook()
    
    # ABA 1: Diagnóstico
    ws1 = wb.active
    ws1.title = "Diagnóstico"
    
    # Seção 1: Visão Geral
    ws1['A1'] = "Seção 1: Visão Geral"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Pergunta 1
    ws1['A2'] = "Qual é o desafio principal da sua empresa?"
    ws1['A2'].font = Font(bold=True)
    ws1.merge_cells('B2:D4')
    ws1['B2'] = "Resposta esperada aqui"
    ws1['B2'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Pergunta 2
    ws1['A5'] = "Descreva seu produto em 3 linhas"
    ws1['A5'].font = Font(bold=True)
    ws1.merge_cells('B5:D8')
    ws1['B5'] = "Resposta"
    ws1['B5'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # ABA 2: Mercado
    ws2 = wb.create_sheet("Mercado")
    
    # Seção 2: TAM
    ws2['A1'] = "Seção 2: Tamanho de Mercado"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Pergunta 3
    ws2['A2'] = "Qual é o TAM (Total Addressable Market) em dólares?"
    ws2['A2'].font = Font(bold=True)
    ws2['B2'] = "0"
    ws2['B2'].number_format = "$#,##0"
    
    # Pergunta 4
    ws2['A3'] = "Qual é o SAM (Serviceable Addressable Market)?"
    ws2['A3'].font = Font(bold=True)
    ws2['B3'] = "0"
    ws2['B3'].number_format = "$#,##0"
    
    # ABA 3: Estratégia
    ws3 = wb.create_sheet("Estratégia")
    
    # Pergunta 5
    ws3['A1'] = "Qual é sua estratégia de go-to-market?"
    ws3['A1'].font = Font(bold=True)
    ws3.merge_cells('B1:D5')
    ws3['B1'] = "Resposta"
    
    # Converter
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# TESTES - FIDELIDADE DE ORDEM
# ============================================================

def test_trail_order_sheets_preserved(trail_workbook_bytes):
    """Verifica que ordem das abas é preservada"""
    
    service = TrailIngestionService()
    questions, report = service.ingest(trail_workbook_bytes)
    
    # Verificar que sheets estão em ordem
    sheet_indices = [q.sheet_index for q in questions]
    assert sheet_indices == sorted(sheet_indices), "Sheet indices não estão em ordem"
    
    # Verificar que sheet 0 vem antes de sheet 1, que vem antes de sheet 2
    sheet_0_questions = [q for q in questions if q.sheet_index == 0]
    sheet_1_questions = [q for q in questions if q.sheet_index == 1]
    sheet_2_questions = [q for q in questions if q.sheet_index == 2]
    
    assert len(sheet_0_questions) > 0
    assert len(sheet_1_questions) > 0
    assert len(sheet_2_questions) > 0
    
    # Verificar que última pergunta de sheet 0 vem antes da primeira de sheet 1
    last_q0 = max(sheet_0_questions, key=lambda q: q.order_index_global)
    first_q1 = min(sheet_1_questions, key=lambda q: q.order_index_global)
    assert last_q0.order_index_global < first_q1.order_index_global


def test_trail_order_questions_within_sheet(trail_workbook_bytes):
    """Verifica que ordem de perguntas dentro da sheet é preservada"""
    
    service = TrailIngestionService()
    questions, report = service.ingest(trail_workbook_bytes)
    
    # Dentro da sheet 0, perguntas devem estar ordenadas por linha
    sheet_0_qs = [q for q in questions if q.sheet_index == 0]
    
    rows = [q.row for q in sheet_0_qs]
    assert rows == sorted(rows), "Perguntas não estão em ordem vertical dentro da sheet"


def test_trail_no_questions_lost(trail_workbook_bytes):
    """Verifica que nenhuma pergunta é perdida"""
    
    service = TrailIngestionService()
    questions, report = service.ingest(trail_workbook_bytes)
    
    # Deve ter pelo menos 5 perguntas (conforme worksheet criado)
    assert len(questions) >= 5, f"Esperava >= 5 perguntas, tem {len(questions)}"
    
    # Todas as perguntas devem ter field_id único
    field_ids = [q.field_id for q in questions]
    assert len(field_ids) == len(set(field_ids)), "Existem field_ids duplicados"


def test_trail_field_id_stable(trail_workbook_bytes):
    """Verifica que field_id é determinístico"""
    
    service = TrailIngestionService()
    questions1, _ = service.ingest(trail_workbook_bytes)
    questions2, _ = service.ingest(trail_workbook_bytes)
    
    # Mesmo arquivo → mesmos field_ids na mesma ordem
    ids1 = [q.field_id for q in questions1]
    ids2 = [q.field_id for q in questions2]
    
    assert ids1 == ids2, "field_ids não são estáveis entre ingestões"


def test_trail_order_index_global_sequential(trail_workbook_bytes):
    """Verifica que order_index_global é sequencial (0, 1, 2, ...)"""
    
    service = TrailIngestionService()
    questions, report = service.ingest(trail_workbook_bytes)
    
    for i, q in enumerate(questions):
        assert q.order_index_global == i, \
            f"Pergunta {i} tem order_index_global={q.order_index_global}"


def test_trail_order_index_sheet_sequential(trail_workbook_bytes):
    """Verifica que order_index_sheet é sequencial dentro de cada sheet"""
    
    service = TrailIngestionService()
    questions, report = service.ingest(trail_workbook_bytes)
    
    # Agrupar por sheet
    by_sheet = {}
    for q in questions:
        if q.sheet_index not in by_sheet:
            by_sheet[q.sheet_index] = []
        by_sheet[q.sheet_index].append(q)
    
    # Verificar sequência dentro de cada sheet
    for sheet_idx, sheet_qs in by_sheet.items():
        indices = [q.order_index_sheet for q in sheet_qs]
        expected = list(range(len(indices)))
        assert indices == expected, \
            f"Sheet {sheet_idx} tem order_index_sheet não sequencial: {indices}"


def test_trail_extraction_audit(trail_workbook_bytes):
    """Verifica que relatório de extração é completo"""
    
    service = TrailIngestionService()
    questions, report = service.ingest(trail_workbook_bytes)
    
    # Verificar estrutura do report
    assert "step_1_snapshot" in report
    assert "step_2_questions" in report
    assert "step_3_validation" in report
    
    # Verificar status
    assert report["step_1_snapshot"]["status"] == "✅ OK"
    assert report["step_2_questions"]["status"] == "✅ OK"
    assert report["step_3_validation"]["status"] == "✅ OK"


def test_trail_section_assignment(trail_workbook_bytes):
    """Verifica que perguntas são associadas a seções"""
    
    service = TrailIngestionService()
    questions, report = service.ingest(trail_workbook_bytes)
    
    # Verificar que seções foram atribuídas
    sections_found = set()
    for q in questions:
        if q.section_name:
            sections_found.add(q.section_name)
    
    # Deve ter pelo menos 1 seção
    assert len(sections_found) > 0, "Nenhuma seção foi atribuída"


def test_trail_reproducibility(trail_workbook_bytes):
    """Verifica que trilha é reprodutível (mesmo resultado em múltiplas ingestões)"""
    
    service = TrailIngestionService()
    
    # Ingerir 3 vezes
    questions_list = []
    for i in range(3):
        qs, _ = service.ingest(trail_workbook_bytes)
        questions_list.append(qs)
    
    # Comparar
    for i in range(1, len(questions_list)):
        q1 = questions_list[0]
        q2 = questions_list[i]
        
        assert len(q1) == len(q2), f"Ingestão {i} teve tamanho diferente"
        
        for j, (qa, qb) in enumerate(zip(q1, q2)):
            assert qa.field_id == qb.field_id, f"Pergunta {j}: field_id diferente"
            assert qa.question_text == qb.question_text, f"Pergunta {j}: texto diferente"
            assert qa.order_index_global == qb.order_index_global, f"Pergunta {j}: ordem diferente"


def test_trail_coverage_validation(trail_workbook_bytes):
    """Verifica que validação de cobertura detecta inconsistências"""
    
    # Criar workbook SEM perguntas em uma aba
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    ws['A1'] = "Apenas dados"
    ws['B1'] = "Sem perguntas"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    empty_bytes = output.getvalue()
    
    service = TrailIngestionService()
    
    # Deve falhar com TrailIngestionError
    with pytest.raises(TrailIngestionError):
        service.ingest(empty_bytes)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
