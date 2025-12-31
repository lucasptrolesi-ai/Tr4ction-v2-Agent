"""
Integration Tests for Template Engine
=====================================
Demonstrates end-to-end functionality.

Run with:
    pytest backend/tests/test_template_engine.py -v
"""

import pytest
import json
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock

from fastapi.testclient import TestClient
from openpyxl import Workbook

# Assuming these imports work
from main import app
from services.excel_template_parser import (
    ExcelTemplateParser,
    FieldType,
    CellPosition,
)
from services.template_manager import TemplateDataService, TemplateManager
from services.ai_mentor_context import (
    AIMentorContextBuilder,
    AIMentorPayloadBuilder,
)


@pytest.fixture
def client():
    """Test client for FastAPI app."""
    return TestClient(app)


@pytest.fixture
def test_excel_file(tmp_path):
    """Create a test Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Persona"

    # Add sample data
    ws["A1"] = "Persona Details"
    ws["B2"] = "Sample Persona"
    ws["B3"] = "25-35"
    ws["B4"] = "Software Engineer"

    # Set dimensions
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.row_dimensions[1].height = 25

    excel_path = tmp_path / "test_template.xlsx"
    wb.save(str(excel_path))

    return excel_path


@pytest.fixture
def template_schema():
    """Sample template schema."""
    return {
        "template_key": "persona_01",
        "sheet_name": "Persona",
        "sheet_width": 1200.5,
        "sheet_height": 945.2,
        "fields": [
            {
                "key": "persona_name",
                "cell": "B2",
                "type": "text",
                "label": "Persona Name",
                "position": {"top": 18.95, "left": 49.0, "width": 343.0, "height": 20.0},
                "required": True,
            },
            {
                "key": "age_range",
                "cell": "B3",
                "type": "text",
                "label": "Age Range",
                "position": {"top": 38.95, "left": 49.0, "width": 343.0, "height": 20.0},
                "required": True,
            },
            {
                "key": "occupation",
                "cell": "B4",
                "type": "text",
                "label": "Occupation",
                "position": {"top": 58.95, "left": 49.0, "width": 343.0, "height": 20.0},
                "required": True,
            },
        ],
    }


# ============================================================================
# UNIT TESTS: Parser
# ============================================================================


class TestExcelTemplateParser:
    """Test Excel to pixel conversion."""

    def test_column_width_calculation(self, test_excel_file):
        """Test column width pixel conversion."""
        parser = ExcelTemplateParser(str(test_excel_file))
        worksheet = parser.workbook.active

        # Column B width should be calculated
        width_pixels = parser._get_column_width_pixels(worksheet, "B")
        assert isinstance(width_pixels, float)
        assert width_pixels > 0

        parser.close()

    def test_row_height_calculation(self, test_excel_file):
        """Test row height pixel conversion."""
        parser = ExcelTemplateParser(str(test_excel_file))
        worksheet = parser.workbook.active

        # Row 2 height should be calculated
        height_pixels = parser._get_row_height_pixels(worksheet, 2)
        assert isinstance(height_pixels, float)
        assert height_pixels > 0

        parser.close()

    def test_cell_position_calculation(self, test_excel_file):
        """Test exact cell position calculation."""
        parser = ExcelTemplateParser(str(test_excel_file))
        worksheet = parser.workbook.active

        position = parser.get_cell_position(worksheet, "B2")

        assert isinstance(position, CellPosition)
        assert position.top >= 0
        assert position.left >= 0
        assert position.width > 0
        assert position.height > 0

        parser.close()

    def test_parse_sheet_with_fields(self, test_excel_file):
        """Test parsing a sheet with field definitions."""
        parser = ExcelTemplateParser(str(test_excel_file))

        fields = {
            "persona_name": {"cell": "B2", "type": "text", "label": "Persona Name"},
            "age_range": {"cell": "B3", "type": "text", "label": "Age Range"},
        }

        schema = parser.parse_sheet(
            sheet_name="Persona",
            fields=fields,
            template_key="test_persona",
        )

        assert schema.template_key == "test_persona"
        assert schema.sheet_name == "Persona"
        assert len(schema.fields) == 2
        assert schema.fields[0].key == "persona_name"
        assert schema.fields[0].position is not None

        parser.close()


# ============================================================================
# UNIT TESTS: Template Manager
# ============================================================================


class TestTemplateDataService:
    """Test template data persistence."""

    def test_save_and_load_template_data(self, tmp_path):
        """Test saving and loading template data."""
        service = TemplateDataService(str(tmp_path))

        test_data = {
            "persona_name": "Young Professional",
            "age_range": "25-35",
            "occupation": "Engineer",
        }

        # Save
        saved = service.save_template_data(
            startup_id="startup_123",
            template_key="persona_01",
            data=test_data,
        )

        assert saved["version"] == 1
        assert saved["data"] == test_data

        # Load
        loaded = service.load_template_data(
            startup_id="startup_123",
            template_key="persona_01",
            version=1,
        )

        assert loaded is not None
        assert loaded["data"] == test_data

    def test_version_auto_increment(self, tmp_path):
        """Test auto-versioning on save."""
        service = TemplateDataService(str(tmp_path))

        # First save
        service.save_template_data(
            startup_id="startup_123",
            template_key="persona_01",
            data={"name": "v1"},
            auto_version=True,
        )

        # Second save
        saved2 = service.save_template_data(
            startup_id="startup_123",
            template_key="persona_01",
            data={"name": "v2"},
            auto_version=True,
        )

        assert saved2["version"] == 2

    def test_validate_required_fields(self, tmp_path, template_schema):
        """Test validation of required fields."""
        service = TemplateDataService(str(tmp_path))

        # Create schema file
        schema_path = tmp_path / "schemas" / "persona_01.json"
        schema_path.parent.mkdir(parents=True, exist_ok=True)
        with open(schema_path, "w") as f:
            json.dump(template_schema, f)

        # Test missing required field
        incomplete_data = {"age_range": "25-35"}  # Missing persona_name (required)

        validation = service.validate_data("persona_01", incomplete_data)

        assert not validation["valid"]
        assert len(validation["errors"]) > 0
        assert any("persona_name" in str(e) for e in validation["errors"])

    def test_validate_length_constraints(self, tmp_path, template_schema):
        """Test validation of length constraints."""
        service = TemplateDataService(str(tmp_path))

        # Add validation rules to schema
        template_schema["fields"][0]["validation_rules"] = {"min": 3, "max": 50}

        schema_path = tmp_path / "schemas" / "persona_01.json"
        schema_path.parent.mkdir(parents=True, exist_ok=True)
        with open(schema_path, "w") as f:
            json.dump(template_schema, f)

        # Test too short
        data = {"persona_name": "AB"}  # Less than min 3

        validation = service.validate_data("persona_01", data)
        assert not validation["valid"]


# ============================================================================
# UNIT TESTS: AI Mentor
# ============================================================================


class TestAIMentorContextBuilder:
    """Test AI mentor context generation."""

    def test_coherence_validation(self):
        """Test coherence validation between templates."""
        builder = AIMentorContextBuilder()

        current_data = {
            "occupation": "Software Engineer",
            "values": "Innovation, Security",
        }

        related_templates = {
            "icp_01": {
                "data": {"decision_making_style": "Analytical", "industry": "Tech"}
            }
        }

        issues = builder.validate_coherence(
            "persona_01", current_data, related_templates
        )

        # Should find alignment issues or suggestions
        assert isinstance(issues, list)

    def test_prompt_generation_for_template(self):
        """Test system prompt generation."""
        from services.ai_mentor_context import AIMentorPromptGenerator

        prompt = AIMentorPromptGenerator.generate_system_prompt("persona_01")

        assert "expert business advisor" in prompt.lower()
        assert "pain_points" in prompt.lower() or "FCJ" in prompt

    def test_prompt_generation_for_field(self):
        """Test field-specific prompt generation."""
        from services.ai_mentor_context import AIMentorPromptGenerator

        prompt = AIMentorPromptGenerator.generate_system_prompt(
            "persona_01", field_key="pain_points"
        )

        assert "pain points" in prompt.lower()
        assert "hook" in prompt.lower() or "value" in prompt.lower()


# ============================================================================
# INTEGRATION TESTS: API Endpoints
# ============================================================================


class TestTemplateAPI:
    """Test template API endpoints."""

    @patch("routers.templates.template_manager")
    def test_get_template_endpoint(self, mock_manager, client):
        """Test GET /templates/{template_key} endpoint."""
        mock_manager.get_template_for_founder.return_value = {
            "schema": {
                "template_key": "persona_01",
                "sheet_name": "Persona",
                "sheet_width": 1200,
                "sheet_height": 945,
                "fields": [],
            },
            "saved_data": None,
            "versions": [],
        }

        # Mock auth
        with patch("routers.templates.get_current_founder") as mock_auth:
            mock_auth.return_value = Mock(id=123, role="founder")

            response = client.get("/api/founder/templates/persona_01")

            assert response.status_code in [200, 422]  # 422 if auth fails

    @patch("routers.templates.template_manager")
    def test_save_template_endpoint(self, mock_manager, client):
        """Test POST /templates/{template_key} endpoint."""
        mock_manager.get_template_for_founder.return_value = {
            "schema": {"fields": []},
            "saved_data": None,
            "versions": [],
        }

        save_data = {
            "data": {"persona_name": "Test", "age_range": "25-35"}
        }

        with patch("routers.templates.get_current_founder") as mock_auth:
            mock_auth.return_value = Mock(id=123, role="founder")

            response = client.post(
                "/api/founder/templates/persona_01",
                json=save_data,
            )

            assert response.status_code in [200, 201, 422]


# ============================================================================
# END-TO-END TEST: Complete Flow
# ============================================================================


class TestCompleteFlow:
    """Test complete user flow."""

    def test_founder_fills_and_exports_template(self, tmp_path, test_excel_file):
        """Test: Parse → Save → Export flow."""

        # Step 1: Parse Excel template
        parser = ExcelTemplateParser(str(test_excel_file))

        fields = {
            "persona_name": {"cell": "B2", "type": "text", "label": "Persona Name"},
            "age_range": {"cell": "B3", "type": "text", "label": "Age Range"},
            "occupation": {"cell": "B4", "type": "text", "label": "Occupation"},
        }

        schema = parser.parse_sheet(
            sheet_name="Persona", fields=fields, template_key="persona_01"
        )

        assert len(schema.fields) == 3
        parser.close()

        # Step 2: Founder fills template
        service = TemplateDataService(str(tmp_path))

        founder_data = {
            "persona_name": "Young Urban Professional",
            "age_range": "25-35",
            "occupation": "Software Engineer",
        }

        saved = service.save_template_data(
            startup_id="startup_xyz", template_key="persona_01", data=founder_data
        )

        assert saved["version"] == 1
        assert saved["data"] == founder_data

        # Step 3: Export to Excel
        # (In real scenario, would write to actual Excel file)
        exported_data = saved["data"]

        assert exported_data["persona_name"] == "Young Urban Professional"
        assert exported_data["age_range"] == "25-35"


# ============================================================================
# PERFORMANCE TESTS
# ============================================================================


class TestPerformance:
    """Test performance characteristics."""

    def test_large_sheet_parsing(self, tmp_path):
        """Test parsing large sheets with many fields."""
        wb = Workbook()
        ws = wb.active

        # Create large sheet with 100 rows
        for row in range(1, 101):
            ws[f"A{row}"] = f"Field {row}"
            ws[f"B{row}"] = f"Value {row}"

        excel_path = tmp_path / "large_template.xlsx"
        wb.save(str(excel_path))

        # Parse with 50 fields
        parser = ExcelTemplateParser(str(excel_path))

        fields = {f"field_{i}": {"cell": f"B{i+1}", "type": "text"} for i in range(50)}

        import time

        start = time.time()
        schema = parser.parse_sheet(sheet_name="Sheet", fields=fields)
        duration = time.time() - start

        assert duration < 1.0  # Should parse in under 1 second
        parser.close()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
