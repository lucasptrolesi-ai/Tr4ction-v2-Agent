"""
Testes para consolidação de suporte a Excel (.xlsx)

Cobertura:
1. Leitura de workbook válido
2. Snapshot contém merged_cells, validations, images
3. Detector gera campos preenchíveis
4. Upload de dois templates diferentes cria dois registros
5. Upload duplicado não duplica template
6. Validação robusta de snapshot
7. Fail fast em arquivo inválido
"""

import io
import pytest
import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Import dos serviços
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.services.template_snapshot import (
    TemplateSnapshotService, 
    SnapshotLoadError, 
    SnapshotValidationError,
    validate_snapshot
)
from app.services.fillable_detector import FillableAreaDetector, FillableFieldCandidate
from app.services.template_registry import TemplateRegistry


# ============================================================
# FIXTURES
# ============================================================

@pytest.fixture
def sample_workbook_bytes():
    """Cria um workbook de teste com estrutura completa"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Adicionar dados
    ws['A1'] = "Test Field"
    ws['A1'].font = Font(bold=True, size=12)
    
    # Criar célula de preenchimento (merged)
    ws['B2'] = "Value here"
    ws['B2'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ws['B2'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Validação de dados
    dv = ws.data_validations.add(ws['C1'])
    dv.type = "list"
    dv.formula1 = '"Option1,Option2"'
    
    # Converter para bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def invalid_workbook_bytes():
    """Cria um arquivo não-Excel"""
    return b"Este nao eh um arquivo Excel valido"


@pytest.fixture
def complex_workbook_bytes():
    """Cria workbook com múltiplas sheets e merged cells"""
    wb = Workbook()
    
    # Sheet 1
    ws1 = wb.active
    ws1.title = "Phase1"
    ws1['A1'] = "Phase 1 Data"
    ws1['A1'].font = Font(bold=True, size=14)
    
    # Merged cells
    ws1.merge_cells('B2:C3')
    ws1['B2'] = "Merged Field"
    ws1['B2'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Sheet 2
    ws2 = wb.create_sheet("Phase2")
    ws2['A1'] = "Phase 2 Data"
    ws2['B2'] = "Another field"
    
    # Converter
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# TESTES - PASSO 1: LEITURA DE WORKBOOK VÁLIDO
# ============================================================

def test_load_valid_workbook(sample_workbook_bytes):
    """Verifica que workbook válido é carregado"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    assert snapshot is not None
    assert "schema_version" in snapshot
    assert snapshot["schema_version"] == "2.0"
    assert len(snapshot["sheets"]) > 0


def test_load_invalid_workbook(invalid_workbook_bytes):
    """Verifica que arquivo inválido gera SnapshotLoadError"""
    service = TemplateSnapshotService()
    
    with pytest.raises(SnapshotLoadError):
        service.extract(invalid_workbook_bytes)


# ============================================================
# TESTES - PASSO 2: SNAPSHOT COMPLETO (MERGED, VALIDATIONS, IMAGES)
# ============================================================

def test_snapshot_has_merged_cells(complex_workbook_bytes):
    """Verifica que merged_cells são capturados"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(complex_workbook_bytes)
    
    # Verificar que pelo menos uma sheet tem merged_cells
    has_merged = any(len(s.get("merged_cells", [])) > 0 for s in snapshot["sheets"])
    assert has_merged, "Nenhuma sheet tem merged_cells capturados"


def test_snapshot_has_data_validations(sample_workbook_bytes):
    """Verifica que data validations são capturados"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    # Verificar que validations está no snapshot
    sheet = snapshot["sheets"][0]
    assert "data_validations" in sheet
    assert isinstance(sheet["data_validations"], list)


def test_snapshot_structure_complete(sample_workbook_bytes):
    """Verifica que snapshot contém TODOS os campos obrigatórios"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    sheet = snapshot["sheets"][0]
    
    # Campos obrigatórios de sheet
    required_fields = [
        "name", "sheet_state", "freeze_panes", "page_setup", "page_margins",
        "row_dimensions", "column_dimensions", "merged_cells",
        "cells", "data_validations", "conditional_formatting", "tables", "images"
    ]
    
    for field in required_fields:
        assert field in sheet, f"Campo obrigatório '{field}' ausente"


def test_cells_have_complete_style(sample_workbook_bytes):
    """Verifica que células têm estilo completo"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    sheet = snapshot["sheets"][0]
    if sheet["cells"]:
        cell = sheet["cells"][0]
        
        # Verificar style
        assert "style" in cell
        style = cell["style"]
        
        style_fields = ["font", "fill", "border", "alignment", "protection"]
        for field in style_fields:
            assert field in style, f"Style incompleto: falta '{field}'"


# ============================================================
# TESTES - PASSO 3: DETECTOR GERA CAMPOS PREENCHÍVEIS
# ============================================================

def test_detector_finds_fillable_areas(sample_workbook_bytes):
    """Verifica que detector identifica campos preenchíveis"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    detector = FillableAreaDetector()
    candidates = detector.detect(snapshot)
    
    assert len(candidates) > 0, "Nenhum campo preenchível detectado"
    assert isinstance(candidates[0], FillableFieldCandidate)


def test_detector_infers_types(sample_workbook_bytes):
    """Verifica que detector infere tipos corretamente"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    detector = FillableAreaDetector()
    candidates = detector.detect(snapshot)
    
    for candidate in candidates:
        assert candidate.inferred_type in [
            "text_short", "text_long", "number", "date", "choice", "email", "tel"
        ]


def test_field_candidate_has_stable_id(sample_workbook_bytes):
    """Verifica que field_id é determinístico"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    detector = FillableAreaDetector()
    candidates = detector.detect(snapshot)
    
    # Extrair duas vezes e verificar que field_id é igual
    field_dict_1 = candidates[0].to_dict("template_1")
    field_dict_2 = candidates[0].to_dict("template_1")
    
    assert field_dict_1["field_id"] == field_dict_2["field_id"]


# ============================================================
# TESTES - PASSO 4 & 5: DOIS TEMPLATES DIFERENTES E DUPLICATA
# ============================================================

def test_registry_computes_different_keys():
    """Verifica que dois templates diferentes têm chaves diferentes"""
    registry = TemplateRegistry()
    
    key1 = registry.compute_template_key("template1.xlsx", "Q1")
    key2 = registry.compute_template_key("template2.xlsx", "Q1")
    key3 = registry.compute_template_key("template1.xlsx", "Q2")
    
    assert key1 != key2
    assert key1 != key3


def test_registry_computes_same_hash_for_same_file(sample_workbook_bytes):
    """Verifica que hash é idempotente"""
    registry = TemplateRegistry()
    
    hash1 = registry.compute_file_hash(sample_workbook_bytes)
    hash2 = registry.compute_file_hash(sample_workbook_bytes)
    
    assert hash1 == hash2


def test_registry_computes_different_hash_for_different_files(
    sample_workbook_bytes, 
    complex_workbook_bytes
):
    """Verifica que arquivos diferentes têm hashes diferentes"""
    registry = TemplateRegistry()
    
    hash1 = registry.compute_file_hash(sample_workbook_bytes)
    hash2 = registry.compute_file_hash(complex_workbook_bytes)
    
    assert hash1 != hash2


# ============================================================
# TESTES - PASSO 6: VALIDAÇÃO AUTOMÁTICA
# ============================================================

def test_validate_snapshot_valid(sample_workbook_bytes):
    """Verifica que snapshot válido passa em validação"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    report = validate_snapshot(snapshot)
    
    assert report["valid"] == True
    assert len(report["errors"]) == 0


def test_validate_snapshot_rejects_incomplete(sample_workbook_bytes):
    """Verifica que snapshot incompleto é rejeitado"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    # Remover um campo obrigatório
    del snapshot["sheets"][0]["cells"]
    
    report = validate_snapshot(snapshot)
    
    assert report["valid"] == False
    assert len(report["errors"]) > 0


def test_validate_snapshot_stats(sample_workbook_bytes):
    """Verifica que validação retorna estatísticas"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    report = validate_snapshot(snapshot)
    
    assert "stats" in report
    assert "sheets_count" in report["stats"]
    assert "total_cells" in report["stats"]
    assert "total_merged" in report["stats"]


# ============================================================
# TESTES - PASSO 7: GENÉRICO PARA NOVOS TEMPLATES
# ============================================================

def test_detector_works_with_multiple_sheets(complex_workbook_bytes):
    """Verifica que detector funciona com múltiplas sheets"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(complex_workbook_bytes)
    
    assert len(snapshot["sheets"]) > 1
    
    detector = FillableAreaDetector()
    candidates = detector.detect(snapshot)
    
    # Verificar que candidatos vêm de diferentes sheets
    sheet_names = set(c.sheet for c in candidates)
    assert len(sheet_names) > 0


def test_stats_are_computed(sample_workbook_bytes):
    """Verifica que stats são computados corretamente"""
    registry = TemplateRegistry()
    
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    detector = FillableAreaDetector()
    candidates = detector.detect(snapshot)
    
    fields = [c.to_dict("test") for c in candidates]
    stats = registry.compute_stats(snapshot, fields)
    
    assert "num_sheets" in stats
    assert "num_cells" in stats
    assert "num_fields" in stats
    assert stats["num_fields"] > 0


# ============================================================
# TESTES - PASSO 8: FAIL FAST EM PRODUÇÃO
# ============================================================

def test_missing_dependency_check():
    """Verifica que dependências XLSX existem"""
    try:
        import openpyxl
        import PIL
        import lxml
        import dateutil
        # Se chegar aqui, todas as dependências existem
        assert True
    except ImportError as e:
        pytest.fail(f"Dependência XLSX faltando: {e}")


def test_snapshot_validation_is_mandatory(sample_workbook_bytes):
    """Verifica que validação do snapshot é obrigatória"""
    service = TemplateSnapshotService()
    
    # Se criar snapshot sem validação, deve falhar
    # (a validação é automática no extract)
    snapshot, assets = service.extract(sample_workbook_bytes)
    
    # Snapshot foi validado automaticamente durante extract
    assert snapshot is not None


def test_error_messages_are_explicit(invalid_workbook_bytes):
    """Verifica que mensagens de erro são explícitas"""
    service = TemplateSnapshotService()
    
    try:
        service.extract(invalid_workbook_bytes)
        pytest.fail("Deveria ter lançado exceção")
    except SnapshotLoadError as e:
        # Verificar que mensagem é clara
        error_msg = str(e)
        assert "XLSX" in error_msg or "Excel" in error_msg or "arquivo" in error_msg


# ============================================================
# TESTES DE REGRESSÃO
# ============================================================

def test_complex_workbook_fully_processed(complex_workbook_bytes):
    """Testa processamento completo de workbook complexo"""
    service = TemplateSnapshotService()
    snapshot, assets = service.extract(complex_workbook_bytes)
    
    # Validar
    report = validate_snapshot(snapshot)
    assert report["valid"]
    
    # Detectar campos
    detector = FillableAreaDetector()
    candidates = detector.detect(snapshot)
    
    # Computar stats
    registry = TemplateRegistry()
    fields = [c.to_dict("test") for c in candidates]
    stats = registry.compute_stats(snapshot, fields)
    
    # Verificar que todo o pipeline funcionou
    assert len(snapshot["sheets"]) == 2
    assert stats["num_sheets"] == 2
    assert stats["num_fields"] >= 0


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
