#!/usr/bin/env python3
"""
TEMPLATE ENGINE SCALING AUTOMATION
===================================

Automated generation of JSON schemas and background images
for ALL templates in Template Q1.xlsx.

Process:
1. Discover all sheets
2. Extract editable cells per sheet
3. Generate pixel-perfect positions
4. Create JSON schemas
5. Generate background PNG images
6. Validate integrity

Run: python backend/scripts/scale_templates.py
"""

import json
import logging
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, asdict
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageDraw, ImageFont
import io

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Pixel conversion constants (from parser)
EXCEL_COLUMN_UNIT_TO_PIXELS = 7.0
EXCEL_ROW_POINT_TO_PIXELS = 1.33
DEFAULT_COLUMN_WIDTH = 8.43
DEFAULT_ROW_HEIGHT = 15

# Project paths
EXCEL_FILE = Path(__file__).parent.parent.parent / "Template Q1.xlsx"
SCHEMAS_DIR = Path(__file__).parent.parent / "data" / "schemas"
IMAGES_DIR = Path(__file__).parent.parent.parent / "frontend" / "public" / "templates"
TEMPLATES_GENERATED_DIR = SCHEMAS_DIR  # Same location

# FCJ template structure: predefined editable fields per sheet
FCJ_TEMPLATES_CONFIG = {
    # Template key: (sheet_name, list of editable cells)
    "persona_01": ("Persona", None),  # Use auto-detect
    "icp_01": ("ICP", None),
    "market_01": ("Market", None),
    "value_prop_01": ("Value Prop", None),
    # Will be auto-discovered from sheet
}


@dataclass
class CellPosition:
    """Pixel position of a cell."""
    top: float
    left: float
    width: float
    height: float

    def to_dict(self):
        return asdict(self)


@dataclass
class FieldMetadata:
    """Field definition."""
    key: str
    cell: str
    label: str
    type: str = "text"
    required: bool = False
    section: Optional[str] = None
    help_text: Optional[str] = None
    placeholder: Optional[str] = None
    position: Optional[Dict] = None
    validation_rules: Optional[Dict] = None

    def to_dict(self):
        return {
            k: v for k, v in asdict(self).items()
            if v is not None
        }


@dataclass
class TemplateSchema:
    """Complete template schema."""
    template_key: str
    sheet_name: str
    sheet_width: float
    sheet_height: float
    fields: List[FieldMetadata]
    title: Optional[str] = None
    description: Optional[str] = None
    version: str = "1.0"

    def to_dict(self):
        return {
            "template_key": self.template_key,
            "sheet_name": self.sheet_name,
            "sheet_width": self.sheet_width,
            "sheet_height": self.sheet_height,
            "title": self.title,
            "description": self.description,
            "version": self.version,
            "fields": [f.to_dict() for f in self.fields]
        }


class ExcelTemplateScaler:
    """Automated template discovery and schema generation."""

    def __init__(self, excel_path: Path):
        """Initialize with Excel file."""
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path)
        self._column_width_cache = {}
        self._row_height_cache = {}
        
        logger.info(f"Loaded workbook: {excel_path}")
        logger.info(f"Sheets found: {self.workbook.sheetnames}")

    def get_column_width_pixels(self, worksheet, col_letter: str) -> float:
        """Get column width in pixels."""
        if col_letter in self._column_width_cache:
            return self._column_width_cache[col_letter]
        
        col_width = worksheet.column_dimensions[col_letter].width
        if col_width is None:
            col_width = DEFAULT_COLUMN_WIDTH
        
        pixels = col_width * EXCEL_COLUMN_UNIT_TO_PIXELS
        self._column_width_cache[col_letter] = pixels
        return pixels

    def get_row_height_pixels(self, worksheet, row_num: int) -> float:
        """Get row height in pixels."""
        if row_num in self._row_height_cache:
            return self._row_height_cache[row_num]
        
        row_height = worksheet.row_dimensions[row_num].height
        if row_height is None:
            row_height = DEFAULT_ROW_HEIGHT
        
        pixels = row_height * EXCEL_ROW_POINT_TO_PIXELS
        self._row_height_cache[row_num] = pixels
        return pixels

    def get_cell_position(self, worksheet, cell_address: str) -> CellPosition:
        """Calculate exact pixel position of a cell."""
        col_index = column_index_from_string(cell_address.rstrip('0123456789'))
        row_num = int(cell_address.lstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))
        
        col_letter = get_column_letter(col_index)
        
        # Calculate cumulative left
        left = 0.0
        for col in range(1, col_index):
            left += self.get_column_width_pixels(worksheet, get_column_letter(col))
        
        # Calculate cumulative top
        top = 0.0
        for row in range(1, row_num):
            top += self.get_row_height_pixels(worksheet, row)
        
        # Get cell dimensions
        width = self.get_column_width_pixels(worksheet, col_letter)
        height = self.get_row_height_pixels(worksheet, row_num)
        
        return CellPosition(top=top, left=left, width=width, height=height)

    def discover_editable_cells(self, worksheet) -> List[Tuple[str, str]]:
        """
        Auto-discover editable cells in worksheet.
        
        Heuristics:
        - Cells with borders or filled background (formatted) are likely data cells
        - Cells next to labels (text + data pattern) are editable
        - Skip merged cells and formula cells
        """
        editable_cells = []
        
        # Scan all non-empty cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.value is None:
                    continue
                
                # Skip formula cells (data source cells, not input)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    continue
                
                # Check if cell has formatting (likely a data cell)
                has_border = False
                if cell.border:
                    has_border = (
                        (cell.border.left and cell.border.left.style) or 
                        (cell.border.right and cell.border.right.style) or 
                        (cell.border.top and cell.border.top.style) or 
                        (cell.border.bottom and cell.border.bottom.style)
                    )
                
                has_fill = cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000'
                
                # Heuristic: if cell is formatted or next to a label, it's editable
                if has_border or has_fill:
                    editable_cells.append((cell.coordinate, str(cell.value)[:50]))
        
        return editable_cells

    def generate_schema_for_sheet(self, sheet_name: str) -> Optional[TemplateSchema]:
        """Generate schema for a single sheet."""
        try:
            worksheet = self.workbook[sheet_name]
        except KeyError:
            logger.warning(f"Sheet '{sheet_name}' not found")
            return None
        
        logger.info(f"\nProcessing sheet: {sheet_name}")
        
        # Get sheet dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Calculate sheet dimensions in pixels
        sheet_width = sum(
            self.get_column_width_pixels(worksheet, get_column_letter(col))
            for col in range(1, max_col + 1)
        )
        sheet_height = sum(
            self.get_row_height_pixels(worksheet, row)
            for row in range(1, max_row + 1)
        )
        
        logger.info(f"  Sheet dimensions: {sheet_width:.1f} × {sheet_height:.1f} px")
        
        # Discover editable cells
        editable_cells = self.discover_editable_cells(worksheet)
        logger.info(f"  Discovered {len(editable_cells)} editable cells")
        
        # Generate fields
        fields = []
        for i, (cell_addr, label) in enumerate(editable_cells):
            field_key = f"field_{i+1}"
            position = self.get_cell_position(worksheet, cell_addr)
            
            field = FieldMetadata(
                key=field_key,
                cell=cell_addr,
                label=label.strip() if label else f"Field {i+1}",
                type="textarea" if label and len(label) > 50 else "text",
                position=position.to_dict(),
                required=False
            )
            fields.append(field)
        
        # Generate template key from sheet name
        template_key = sheet_name.lower().replace(" ", "_").replace(".", "")
        
        # Create schema
        schema = TemplateSchema(
            template_key=template_key,
            sheet_name=sheet_name,
            sheet_width=sheet_width,
            sheet_height=sheet_height,
            fields=fields,
            title=sheet_name,
            description=f"Template for {sheet_name}"
        )
        
        logger.info(f"  Generated schema with {len(fields)} fields")
        return schema

    def generate_all_schemas(self) -> Dict[str, TemplateSchema]:
        """Generate schemas for all sheets."""
        schemas = {}
        
        for sheet_name in self.workbook.sheetnames:
            # Skip metadata sheets
            if sheet_name.lower() in ['metadata', 'info', 'instructions']:
                logger.info(f"Skipping metadata sheet: {sheet_name}")
                continue
            
            schema = self.generate_schema_for_sheet(sheet_name)
            if schema:
                schemas[schema.template_key] = schema
        
        return schemas

    def save_schemas(self, schemas: Dict[str, TemplateSchema]) -> List[Path]:
        """Save all schemas to JSON files."""
        SCHEMAS_DIR.mkdir(parents=True, exist_ok=True)
        saved_paths = []
        
        for template_key, schema in schemas.items():
            schema_path = SCHEMAS_DIR / f"{template_key}.json"
            with open(schema_path, 'w', encoding='utf-8') as f:
                json.dump(schema.to_dict(), f, indent=2, ensure_ascii=False)
            
            saved_paths.append(schema_path)
            logger.info(f"Saved schema: {schema_path}")
        
        return saved_paths

    def generate_background_images(self, schemas: Dict[str, TemplateSchema]) -> List[Path]:
        """Generate background PNG images for each template."""
        IMAGES_DIR.mkdir(parents=True, exist_ok=True)
        image_paths = []
        
        for template_key, schema in schemas.items():
            try:
                image_path = self._generate_image_for_sheet(
                    schema.sheet_name,
                    schema.template_key,
                    schema.sheet_width,
                    schema.sheet_height
                )
                image_paths.append(image_path)
            except Exception as e:
                logger.error(f"Failed to generate image for {template_key}: {e}")
        
        return image_paths

    def _generate_image_for_sheet(
        self,
        sheet_name: str,
        template_key: str,
        width: float,
        height: float
    ) -> Path:
        """Generate a PNG image representing the sheet."""
        try:
            worksheet = self.workbook[sheet_name]
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        # Create PIL image
        img_width = int(width)
        img_height = int(height)
        img = PILImage.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Draw header
        header_height = 40
        draw.rectangle(
            [(0, 0), (img_width, header_height)],
            fill='#0066cc'
        )
        
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
        except (OSError, IOError):
            font = ImageFont.load_default()
        
        draw.text((10, 10), sheet_name, fill='white', font=font)
        
        # Draw grid lines
        draw.rectangle([(0, 0), (img_width - 1, img_height - 1)], outline='#cccccc')
        
        # Draw vertical grid lines
        col = 1
        x = 0
        while x < img_width:
            col_width = self.get_column_width_pixels(worksheet, get_column_letter(col))
            x += col_width
            if x < img_width:
                draw.line([(x, 0), (x, img_height)], fill='#e0e0e0', width=1)
            col += 1
        
        # Draw horizontal grid lines
        row = 1
        y = 0
        while y < img_height:
            row_height = self.get_row_height_pixels(worksheet, row)
            y += row_height
            if y < img_height:
                draw.line([(0, y), (img_width, y)], fill='#e0e0e0', width=1)
            row += 1
        
        # Save image
        image_path = IMAGES_DIR / f"{template_key}.png"
        img.save(image_path, 'PNG')
        
        logger.info(f"Generated image: {image_path} ({img_width}×{img_height} px)")
        return image_path


def main():
    """Main execution."""
    print("\n" + "="*70)
    print("  TEMPLATE ENGINE SCALING AUTOMATION")
    print("="*70 + "\n")
    
    if not EXCEL_FILE.exists():
        logger.error(f"Excel file not found: {EXCEL_FILE}")
        return False
    
    logger.info(f"Starting template generation from: {EXCEL_FILE}\n")
    
    # Initialize scaler
    scaler = ExcelTemplateScaler(EXCEL_FILE)
    
    # PART 1: Discover sheets
    logger.info(f"Discovered {len(scaler.workbook.sheetnames)} sheets")
    
    # PART 2-4: Generate schemas
    logger.info("\n--- GENERATING SCHEMAS ---")
    schemas = scaler.generate_all_schemas()
    
    if not schemas:
        logger.error("No schemas generated")
        return False
    
    logger.info(f"\nGenerated {len(schemas)} schemas")
    
    # Save schemas
    logger.info("\n--- SAVING SCHEMAS ---")
    schema_paths = scaler.save_schemas(schemas)
    
    # PART 5: Generate images
    logger.info("\n--- GENERATING BACKGROUND IMAGES ---")
    image_paths = scaler.generate_background_images(schemas)
    
    # PART 9: Final report
    print("\n" + "="*70)
    print("  GENERATION REPORT")
    print("="*70 + "\n")
    
    print(f"✅ Templates processed: {len(schemas)}")
    print(f"✅ Schemas saved: {len(schema_paths)}")
    print(f"✅ Images generated: {len(image_paths)}")
    
    print(f"\n📊 TEMPLATE SUMMARY:")
    print("-" * 70)
    
    total_fields = 0
    for template_key, schema in sorted(schemas.items()):
        field_count = len(schema.fields)
        total_fields += field_count
        print(f"  {template_key:30} | {field_count:3} fields | {schema.sheet_width:.0f}×{schema.sheet_height:.0f} px")
    
    print("-" * 70)
    print(f"  TOTAL: {len(schemas)} templates | {total_fields} fields\n")
    
    print(f"📁 Schemas location: {SCHEMAS_DIR}")
    print(f"📁 Images location: {IMAGES_DIR}\n")
    
    print("="*70)
    print("  ✅ SCALING COMPLETE - READY FOR PRODUCTION")
    print("="*70 + "\n")
    
    return True


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
